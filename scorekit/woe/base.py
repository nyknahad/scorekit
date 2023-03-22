# -*- coding: utf-8 -*-

from ..data import DataSamples
from .._utils import color_background, add_suffix, rem_suffix, is_cross, cross_name, cross_split
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.model_selection import cross_val_score, StratifiedKFold, GridSearchCV, PredefinedSplit
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
import re
import os
import openpyxl
import datetime
import gc
import copy
import itertools
import warnings
from tqdm import tqdm
from textwrap import wrap
from concurrent import futures
from functools import partial
import json
try:
    import optbinning
except:
    print('Warning! Optbinning not found.')

warnings.simplefilter('ignore')
plt.rc('font', family='Verdana', size=12)
plt.style.use('seaborn-darkgrid')
pd.set_option('display.precision', 3)
gc.enable()


class WOE:
    """
    Класс для ВОЕ-трансформации переменных
    """
    def __init__(self, ds=None, features=None, scorecard=None, round_digits=3, rounding_migration_coef=0.001,
                 simple=True, n_folds=5, woe_adjust=0.5, alpha=0, alpha_range=None, alpha_scoring='neg_log_loss',
                 alpha_best_criterion='min', missing_process='nearest_or_separate', missing_min_part=0.01,
                 others='missing_or_min', opposite_sign_to_others=False):
        """
        :param ds: ДатаСэмпл, для которого будут рассчитываться биннинги
        :param features: список переменных. При None берется ds.features
        :param scorecard: путь к эксель файлу или датафрейм с готовыми биннингами для импорта

        ---Параметры для расчета WOE---
        :param round_digits: число знаков после запятой для округление границ бинов.
                 При округлении границ бинов происходит проверка на долю мигрирующих наблюдений. Если округление приедет к миграции большой доли наблюдений,
                 то round_digits увеличивается до тех пор, пока доля не упадет ниже rounding_migration_coef
        :param rounding_migration_coef: максимально допустимая доля наблюдений для миграции между бинами при округлении
        :param simple: если True, то расчет WOE происходит на трэйн сэмпле, иначе берется среднее значение по фолдам
        :param n_folds: кол-во фолдов для расчета WOE при simple=False
        :param woe_adjust: корректировочный параметр для расчета EventRate_i в бине i:
                           EventRate_i = (n1_i+woe_adjust)/(n0_i+woe_adjust),
                           где n1_i - кол-во наблюдений с target=1 в бине i, n0_i - кол-во наблюдений с target=0 в бине i
        :param alpha: коэффициент регуляризации для расчета WOE. Если задан, то WOE для бина i вычисляется по формуле:
                      SmoothedWOE_i = log((n + alpha)*EventRate/(n*EventRate_i + alpha)),
                      где n - число наблюдений, EventRate = n1/n0,
                      n1 - общее кол-во наблюдений с target=1, n0 - общее кол-во наблюдений с target=0
        :param alpha_range: если alpha=None, то подбирается оптимальное значение alpha из диапазона alpha_range. При None берется диапазон range(10, 100, 10)
        :param alpha_scoring: метрика, используемая для оптимизации alpha
        :param alpha_best_criterion: 'min' - минимизация метрики alpha_scoring, 'max' - максимизация метрики

        ---Обработка пустых значений---
        :param missing_process: способ обработки пустых значений
                'separate' - помещать в отдельный бин
                'min' - объединять с бином с минимальным WOE
                'max' - объединять с бином с максимальным WOE
                'nearest' - объединять с ближайшим по WOE биномом
                'min_or_separate' - если доля пустых значений меньше missing_min_part, то объединять с худшим бином, иначе помещать в отдельный бин
                'max_or_separate' - если доля пустых значений меньше missing_min_part, то объединять с лучшим бином, иначе помещать в отдельный бин
                'nearest_or_separate' - если доля пустых значений меньше missing_min_part, то объединять с ближайшим по WOE бином, иначе помещать в отдельный бин
        :param missing_min_part: минимальная доля пустых значений для выделения отдельного бина при missing_process 'min_or_separate', 'max_or_separate' или 'nearest_or_separate'

        ---Обработка остальных значений, не попавших в биннинг---
        :param others: Способ обработки значений, не попавших в биннинг:
                'min': остальным значениям присваивается минимальный WOE
                'max': остальным значениям присваивается максимальный WOE
                'missing_or_min': если есть бакет с пустыми значениями, то остальным значениям присваивается его WOE, иначе минимальный WOE
                'missing_or_max': если есть бакет с пустыми значениями, то остальным значениям присваивается его WOE, иначе максимальный WOE
                float: отсутствующим значениям присваивается заданный фиксированный WOE
        :param opposite_sign_to_others: В случае, когда непрерывная переменная на выборке для разработки имеет только один знак,
                то все значения с противоположным знаком относить к others
        """
        self.round_digits = round_digits
        self.round_woe = 3
        self.rounding_migration_coef = rounding_migration_coef
        self.simple = simple
        self.n_folds = n_folds
        self.woe_adjust = woe_adjust
        self.alpha = alpha
        self.alpha_range = alpha_range if self.alpha is None else None
        self.alpha_scoring = alpha_scoring
        self.alpha_best_criterion = alpha_best_criterion
        self.missing_process = missing_process
        self.missing_min_part = missing_min_part
        if others in ['min', 'max', 'missing_or_min', 'missing_or_max'] or isinstance(others, (int, float)):
            self.others = others
        else:
            print('WARNING! Parameter others is incorrect. Set others = "missing_or_min".')
            self.others = 'missing_or_min'

        self.opposite_sign_to_others = opposite_sign_to_others

        if ds is None:
            ds = DataSamples()
        self.special_bins = {} if ds.special_bins is None else ds.special_bins
        if isinstance(scorecard, str):
            scorecard = pd.read_excel(scorecard)
        if features is None:
            if ds.features:
                features = ds.features
            elif scorecard is not None:
                features = list(scorecard['feature'].unique())
        self.feature_woes = {rem_suffix(f): self.create_feature_woe(ds, rem_suffix(f)) for f in features if not is_cross(f)}
        self.feature_crosses = {}

        if ds.samples is not None and (ds.time_column is not None or ds.id_column is not None):
            samples = {name: ds.samples[name][[f for f in [ds.time_column, ds.id_column] if f is not None]] for name in ds.samples}
        else:
            samples = None
        self.ds_aux = DataSamples(samples=samples, features=[], cat_columns=[], time_column=ds.time_column, id_column=ds.id_column,
                                  feature_descriptions=ds.feature_descriptions, train_name=ds.train_name,
                                  result_folder=ds.result_folder, n_jobs=ds.n_jobs)
        if ds.bootstrap_base is not None and ds.time_column is not None:
            self.ds_aux.bootstrap_base = ds.bootstrap_base[[ds.time_column]]
        if scorecard is not None:
            self.import_scorecard(scorecard, verbose=False, fit_flag=False)

    def create_feature_woe(self, ds, f):
        """
        Создание объекта класса FeatureWOE
        :param ds: ДатаСэмпл для обработки
        :param f: переменная

        :return: FeatureWOE
        """
        return FeatureWOE(ds, f, round_digits=self.round_digits, round_woe=self.round_woe,
                          rounding_migration_coef=self.rounding_migration_coef, simple=self.simple, n_folds=self.n_folds,
                          woe_adjust=self.woe_adjust, alpha=self.alpha, alpha_range=self.alpha_range,
                          alpha_scoring=self.alpha_scoring, alpha_best_criterion=self.alpha_best_criterion,
                          missing_process=self.missing_process, missing_min_part=self.missing_min_part,
                          others_process=self.others, opposite_sign_to_others=self.opposite_sign_to_others,
                          special_bins=self.special_bins)

    def fit(self, features=None, new_groups=True, plot_flag=True, method='tree', max_n_bins=10, min_bin_size=0.05,
            criterion='entropy', scoring='neg_log_loss', max_depth=None, monotonic=False, solver='cp', divergence='iv'):
        """
        Пересчет биннинга для списка переменных
        :param features: список переменных для обработки. При None обрабатываются list(self.feature_woes) + list(self.cross_features)
        :param new_groups: False - пересчитаются только WOE в биних,
                           True - также пересчитываются и границы бинов
        :param plot_flag: флаг для вывода графиков с биннингом

        --- Метод биннинга ---
        :param method: 'tree' - биннинг деревом, 'opt' - биннинг деревом с последующей оптимизацией границ бинов библиотекой optbinning
        :param max_n_bins: максимальное кол-во бинов
        :param min_bin_size: минимальное число (доля) наблюдений в каждом листе дерева.
                                Если min_bin_size < 1, то трактуется как доля наблюдений от обучающей выборки

        --- Параметры биннинга для метода 'tree' ---
        :param criterion: критерий расщепления. Варианты значений: 'entropy', 'gini'
        :param scoring: метрика для оптимизации
        :param max_depth: максимальная глубина дерева

        --- Параметры биннинга для метода 'opt' ---
        :param monotonic: флаг для оптимизации биннинга к монотонному тренду
        :param solver: солвер для оптимизации биннинга:
                        'cp' - constrained programming
                        'mip' - mixed-integer programming
                        'ls' - LocalSorver (www.localsorver.com)
        :param divergence: метрика для максимизации:
                        'iv' - Information Value,
                        'js' - Jensen-Shannon,
                        'hellinger' - Hellinger divergence,
                        'triangular' - triangular discrimination
        """
        cross_features = self.cross_features
        if features is None:
            features = list(self.feature_woes) + list(cross_features)
        for f in features:
            if f in self.feature_woes:
                self.feature_woes[f].fit(new_groups=new_groups, method=method, max_n_bins=max_n_bins,
                                         min_bin_size=min_bin_size, criterion=criterion, scoring=scoring,
                                         max_depth=max_depth, monotonic=monotonic, solver=solver,
                                         divergence=divergence)
            elif f in cross_features:
                f1, f2 = cross_split(f)
                self.feature_crosses[f1].fit(cross_features=[f2], new_groups=new_groups, method=method, max_n_bins=max_n_bins,
                                         min_bin_size=min_bin_size, criterion=criterion, scoring=scoring,
                                         max_depth=max_depth, monotonic=monotonic, solver=solver,
                                         divergence=divergence)
        if plot_flag:
            self.plot_bins(features=features, folder=None, plot_flag=True, show_groups=True)

    @staticmethod
    def get_f2_impurity(feature_woe, df1_group, criterion='entropy', max_depth=5, min_bin_size=0.05, max_n_bins=10):
        impurity = 0
        min_samples_leaf = min_bin_size if min_bin_size > 1 else int(round(len(df1_group) * min_bin_size, 0))
        for group, df in feature_woe.ds.samples[feature_woe.ds.train_name].drop(['group'], axis=1) \
                .merge(df1_group, how='left', left_index=True, right_index=True).groupby('group'):
            df = df[~df[feature_woe.feature].isnull()]
            if len(df) > min_samples_leaf* 2:
                x_train = df[feature_woe.feature]
                if feature_woe.categorical_type:
                    groups = {group: [value] for group, value in enumerate(x_train.dropna().unique())}
                    df['group'] = feature_woe.set_groups(data=x_train, groups=groups, inplace=False)
                    x_train = df['group'].map(feature_woe.calc_simple_woes(df=df))
                y_tran = df[feature_woe.ds.target]
                try:
                    tree = DecisionTreeClassifier(criterion=criterion, max_depth=max_depth, min_samples_leaf=min_samples_leaf,
                                                  max_leaf_nodes=max_n_bins,
                                                  random_state=feature_woe.ds.random_state)
                    tree.fit(x_train[:, None], y_tran)
                    impurity += sum(tree.tree_.impurity)
                except:
                    pass
        return impurity

    @staticmethod
    def auto_fit_feature(feature_obj, auto_fit_parameters, verbose):
        """
        Автобиннинг одной переменной
        :param feature_obj: объект класса FeatureWOE или FeatureCross
        :param auto_fit_parameters: словарь с параметрами атвобиннинга
        :param verbose: флаг для вывода подробных комментариев в процессе работы

        :return: флаг успешного завершения, список из датафреймов с логами
        """
        return feature_obj.auto_fit(verbose=verbose, **auto_fit_parameters)

    def auto_fit(self, features=None, autofit_folder='auto_fit', plot_flag=-1, verbose=False,
                 params_space=None, woe_best_samples=None, method='opt', max_n_bins=10, min_bin_size=0.05,
                 criterion='entropy', scoring='neg_log_loss', max_depth=5, solver='cp', divergence='iv',
                 WOEM_on=True, WOEM_woe_threshold=0.05, WOEM_with_missing=False,
                 SM_on=False, SM_target_threshold=5, SM_size_threshold=100,
                 BL_on=True, BL_allow_Vlogic_to_increase_gini=15,
                 G_on=False, G_gini_threshold=5, G_with_test=False, G_gini_decrease_threshold=0.2, G_gini_increase_restrict=True,
                 WOEO_on=False, WOEO_dr_threshold=0.01, WOEO_correct_threshold=0.85, WOEO_miss_is_incorrect=True, WOEO_with_test=False,
                 cross_features_first_level=None, cross_num_second_level=0):
        """
        Поиск оптимального биннинга, удовлетворяющего набору проверок. Итерационно для каждой переменной выполняются следующие шаги:
            1) Исходное разбиение на n бинов, где n на первой итерации равно max_n_bins
            2) Выполняются слияния соседних бинов с близким WOE (при выставленном флаге WOEM_on=True) и малых бинов (SM_on=True)
            3) Проводятся проверки на бизнес-логику (BL_on=True), джини (G_on=True) и порядок (WOEO_on=True)
            4) Если любая из проверок проваливается, то уменьшаем n на 1 и возвращаемся на шаг 1
        Если после перебора всех n проверки так и не успешны, то переменная исключается из списка.
        В случае, если задан params_space, то вся процедура повторяется для каждого набора параметров и
        затем выбирается биннинг с наибольшим джини.

        :param features: список переменных для обработки. При None обрабатываются все self.feature_woes
        :param autofit_folder: название папки, в которую будут сохранены результаты автобиннинга
        :param plot_flag: флаг для вывода графиков с биннингом:
                            -1 - графики не строить
                            0, False - графики сохранить в папку autofit_folder/Figs_binning, но не выводить в аутпут
                            1, True - графики сохранить в папку autofit_folder/Figs_binning и вывести в аутпут
        :param verbose: флаг для вывода подробных комментариев в процессе работы

        --- Метод биннинга ---
        :param method: 'tree' - биннинг деревом, 'opt' - биннинг деревом с последующей оптимизацией границ бинов библиотекой optbinning
        :param max_n_bins: максимальное кол-во бинов
        :param min_bin_size: минимальное число (доля) наблюдений в каждом листе дерева.
                                Если min_bin_size < 1, то трактуется как доля наблюдений от обучающей выборки

        --- Параметры биннинга для метода 'tree' ---
        :param criterion: критерий расщепления. Варианты значений: 'entropy', 'gini', 'log_loss'
        :param scoring: метрика для оптимизации
        :param max_depth: максимальная глубина дерева

        --- Параметры биннинга для метода 'opt' ---
        :param solver: солвер для оптимизации биннинга:
                        'cp' - constrained programming
                        'mip' - mixed-integer programming
                        'ls' - LocalSorver (www.localsorver.com)
        :param divergence: метрика для максимизации:
                        'iv' - Information Value,
                        'js' - Jensen-Shannon,
                        'hellinger' - Hellinger divergence,
                        'triangular' - triangular discrimination

        --- Параметры проверок ---
        :param WOEM_on: флаг проверки на разницу WOE между соседними бинами
        :param WOEM_woe_threshold: минимальная разрешенная дельта WOE между соседними бинами
        :param WOEM_with_missing: должна ли выполняться проверка для бина с пустыми значениями
        :param SM_on: флаг проверки на размер бина
        :param SM_target_threshold: минимальное кол-во (доля) наблюдений с целевым событием в бине
        :param SM_size_threshold: минимальное кол-во (доля) наблюдений в бине
        :param BL_on: флаг проверки на бизнес-логику
        :param BL_allow_Vlogic_to_increase_gini: разрешить V-образную бизнес-логику, если она приводит к увеличению джини переменной на эту величину относительного монотонного тренда.
                                                 При значении 100 V-образная бизнес-логика запрещена
        :param G_on: флаг проверки на джини
        :param G_gini_threshold: минимальное допустимое джини переменной.
                Проверяется на трэйне + если заданы бутстрэп сэмплы, то проверяется на них условие mean-1.96*std > G_gini_threshold
        :param G_with_test: так же проверяется джини на всех остальных доступных сэмплах.
        :param G_gini_decrease_threshold: допустимое уменьшение джини на всех сэмплах относительно трэйна.
                В случае, если значение >= 1, то проверяется условие gini(train) - gini(sample) <= G_gini_decrease_threshold для основных сэмплов
                                                                   и 1.96*std <= G_gini_decrease_threshold для бутсрэп сэмплов
                          если значение < 1, то проверяется условие 1 - gini(sample)/gini(train) <= G_gini_decrease_threshold для основных сэмплов
                                                                  и 1.96*std/mean <= G_gini_decrease_threshold для бутсрэп сэмплов
        :param G_gini_increase_restrict: такое же ограничение действует и на увеличение джини
        :param WOEO_on: флаг проверки на сохранение тренда WOE на бутстрэп-сэмплах
        :param WOEO_dr_threshold: допустимая дельта между TargetRate соседних бинов для прохождения проверки, в случае нарушения тренда
        :param WOEO_correct_threshold: доля бутстрэп-сэмплов, на которых должна проходить проверка
        :param WOEO_miss_is_incorrect: считать ли отсутствие данных в бине сэмпла ошибкой или нет
        :param WOEO_with_test: так же проверять тренд на остальных доступных сэмплах.

        --- Пространство параметров ---
        :param params_space: пространство параметров, с которыми будут выполнены автобиннинги.
                Задается в виде словаря {параметр: список значений}
        :param woe_best_samples: список сэмплов, джини которых будет учитываться при выборе лучшего биннинга.  При None берется джини на трэйне

        --- Кросс переменные ---
        :param cross_features_first_level: список переменных первого уровня для которых будут искаться лучшие кросс пары. При None берется features
        :param cross_num_second_level: кол-во кросс пар, рассматриваемых для каждой переменной первого уровня
                                           0 - поиск не производится
                                           -1 - рассматриваются все возможные кросс пары
                                           n - для каждой переменной первого уровня отбираются n лучших переменных с максимальной метрикой criterion
        """

        # feature_log, feature_gini, feature_bl, feature_woe, feature_er, woe_out
        def save_check_dfs(dfs, woe_best_samples, check_file):
            def transparent(x):
                if pd.isnull(x):
                    return 'background-color: transparent'

            def save_res_file(df, i, writer):
                fix_columns = ['bin1', 'feature1', 'feature', 'iteration', 'group', 'Bootstrap std']
                if df.empty:
                    return None
                if i == 0:
                    df.to_excel(writer, sheet_name='Log', index=False)
                    worksheet = writer.sheets['Log']
                    worksheet.column_dimensions['A'].width = 40
                    worksheet.column_dimensions['B'].width = 12
                    worksheet.column_dimensions['C'].width = 12
                    worksheet.column_dimensions['D'].width = 12
                    worksheet.column_dimensions['E'].width = 60
                elif i == 1:
                    df.to_excel(writer, sheet_name='Business Logic', index=False)
                    worksheet = writer.sheets['Business Logic']
                    worksheet.column_dimensions['A'].width = 40
                    for cn in range(2, worksheet.max_column + 1):
                        worksheet.column_dimensions[openpyxl.utils.get_column_letter(cn)].width = 15
                    worksheet.freeze_panes = worksheet['C2']
                elif i == 2:
                    gini_columns = [x for x in df.columns if x not in fix_columns]
                    df.style.apply(color_background, mn=df[gini_columns].min().min(), mx=df[gini_columns].max().max(),
                                   cmap='RdYlGn', subset=pd.IndexSlice[:, gini_columns]).applymap(
                        lambda x: transparent(x)) \
                        .to_excel(writer, sheet_name='Gini by Samples', index=False,
                                  float_format=f'%0.{self.round_woe}f')
                    worksheet = writer.sheets['Gini by Samples']

                    worksheet.column_dimensions['A'].width = 40
                    worksheet.column_dimensions['B'].width = 12
                    for cn in range(3, worksheet.max_column + 1):
                        cl = openpyxl.utils.get_column_letter(cn)
                        worksheet.column_dimensions[cl].width = 15
                    worksheet.freeze_panes = worksheet['C2']
                elif i == 3:
                    woe_er_columns = [x for x in df.columns if x not in fix_columns]
                    woes_values = df[woe_er_columns].values.reshape(-1, ).tolist()
                    df.fillna(0).style.apply(color_background,
                                   mn=np.mean(woes_values) - 2 * np.std(woes_values),
                                   mx=np.mean(woes_values) + 2 * np.std(woes_values),
                                   cmap='RdYlGn',
                                   subset=woe_er_columns).to_excel(writer, sheet_name='WoE by Samples', index=False,
                                                                   float_format=f'%0.{self.round_woe}f')
                    worksheet = writer.sheets['WoE by Samples']
                    worksheet.column_dimensions['A'].width = 40
                    worksheet.column_dimensions['B'].width = 12
                    worksheet.column_dimensions['C'].width = 12
                    for cn in range(4, worksheet.max_column + 1):
                        cl = openpyxl.utils.get_column_letter(cn)
                        worksheet.column_dimensions[cl].width = 12
                    worksheet.freeze_panes = worksheet['D2']
                elif i == 4:
                    woe_er_columns = [x for x in df.columns if x not in fix_columns]
                    er_values = df[woe_er_columns].values.reshape(-1, ).tolist()
                    df.fillna(0).style.apply(color_background,
                                   mn=max([0, np.mean(er_values) - 2 * np.std(er_values)]),
                                   mx=np.mean(er_values) + 2 * np.std(er_values),
                                   cmap='RdYlGn_r',
                                   subset=woe_er_columns).to_excel(writer, sheet_name='Event Rate by Samples', index=False)
                    worksheet = writer.sheets['Event Rate by Samples']
                    worksheet.column_dimensions['A'].width = 40
                    worksheet.column_dimensions['B'].width = 12
                    worksheet.column_dimensions['C'].width = 12
                    for cn in range(4 if 'bin1' not in df.columns else 5, worksheet.max_column + 1):
                        cl = openpyxl.utils.get_column_letter(cn)
                        worksheet.column_dimensions[cl].width = 12
                        for cell in worksheet[cl]:
                            cell.number_format = '0.000%'
                    worksheet.freeze_panes = worksheet['D2']
                elif i == 6:
                    df.to_excel(writer, sheet_name='Params')
                    worksheet = writer.sheets['Params']
                    worksheet.column_dimensions['A'].width = 40
                    worksheet.column_dimensions['B'].width = 20

            f_gini = {}
            Vlogic_features = set()
            if dfs:
                with pd.ExcelWriter(check_file, engine="openpyxl") as writer, \
                     pd.ExcelWriter(check_file[:-5] + '_all.xlsx', engine="openpyxl") as writer_all:
                    for i in range(5):
                        res_dfs = [dfs[f][i] for f in dfs if dfs[f]]
                        if not res_dfs:
                            continue
                        res_df = pd.concat(res_dfs).reset_index(drop=True)
                        if 'feature1' in res_df.columns:
                            group_columns = ['feature1', 'bin1', 'feature']
                        else:
                            group_columns = ['feature']
                        save_res_file(res_df, i, writer_all)
                        if not res_df.empty:
                            if 'iteration' in res_df.columns:
                                res_df = res_df[res_df.groupby(group_columns)['iteration'].transform('max') == res_df['iteration']]
                            save_res_file(res_df, i, writer)
                            if i == 1:
                                Vlogic_features = set(res_df[res_df['trend_type'] == 'V-shape']['feature'].values)
                            if i == 2:
                                #res_df = res_df[res_df['feature'].isin([f for f in features if self.feature_woes[f].is_active])]
                                f_gini = res_df.set_index(group_columns).round(self.round_woe)[[f for f in woe_best_samples if f in res_df.columns]].mean(axis=1).to_dict()
                    df_params = pd.DataFrame().from_dict(auto_fit_parameters, orient='index', columns=['value'])
                    save_res_file(df_params, 6, writer)
                    save_res_file(df_params, 6, writer_all)
            return f_gini, Vlogic_features

        print(f"\n{' SFA ':-^150}\n")
        autofit_folder = autofit_folder.rstrip('/')
        folder = self.ds_aux.result_folder + autofit_folder + '/'
        if not os.path.exists(folder):
            os.makedirs(folder)

        if features is None:
            features = list(self.feature_woes)
        cross_features = [f for f in features if is_cross(f)]
        features = [f for f in features if not is_cross(f)]
        auto_fit_parameters = {}
        for f in ['method', 'max_n_bins', 'min_bin_size', 'criterion', 'scoring', 'max_depth', 'solver', 'divergence',
                  'WOEM_on',  'WOEM_woe_threshold', 'WOEM_with_missing', 'SM_on', 'SM_target_threshold', 'SM_size_threshold',
                  'G_on', 'G_gini_threshold', 'G_gini_decrease_threshold', 'G_gini_increase_restrict', 'G_with_test',
                  'WOEO_on','WOEO_dr_threshold', 'WOEO_correct_threshold',  'WOEO_miss_is_incorrect', 'WOEO_with_test',
                  'BL_on', 'BL_allow_Vlogic_to_increase_gini']:
            auto_fit_parameters[f] = eval(f)
        if params_space is None:
            params_space = {k: [v] for k, v in auto_fit_parameters.items()}
        params_names = list(params_space.keys())
        params_list = list(itertools.product(*params_space.values()))
        file_scorecard = f'{self.ds_aux.result_folder}{autofit_folder}_scorecard.xlsx'
        if woe_best_samples:
            woe_best_samples = [f for f in woe_best_samples if f in self.ds_aux.samples or 'Bootstrap' if f]
        if not woe_best_samples:
            woe_best_samples = [self.ds_aux.train_name]
        scorecard = pd.DataFrame()

        # ------------------------------------------ autobinning for features ------------------------------------------
        if features:
            print(f'Performing autobinning with parameters space of size {len(params_list)}...')
            f_gini = {}
            Vlogic_features = {}
            # check_dfs = {num_i: {feature: [feature_log, feature_bl, feature_gini, feature_woe, feature_er, woe_out]}}
            check_dfs = {num_p: {} for num_p in range(1, len(params_list) + 1)}
            for num_p, param in enumerate(params_list, start=1):
                auto_fit_parameters.update({params_names[k]: param[k] for k in range(len(param))})
                print(f'Using parameters set {num_p}/{len(params_list)}: {auto_fit_parameters}')
                for p in ['SM_target_threshold', 'SM_size_threshold']:
                    if auto_fit_parameters[p] and auto_fit_parameters[p] < 1:
                        auto_fit_parameters[p] = int(round(len(self.ds_aux.samples[self.ds_aux.train_name]) * auto_fit_parameters[p], 0))
                print(f'Processing {len(features)} features on {self.ds_aux.n_jobs} CPU{"s" if self.ds_aux.n_jobs > 1 else ""}...')
                if self.ds_aux.n_jobs > 1 and len(features) > self.ds_aux.n_jobs:
                    with futures.ProcessPoolExecutor(max_workers=self.ds_aux.n_jobs) as pool:
                        for i, result in enumerate(tqdm(pool.map(partial(self.auto_fit_feature, auto_fit_parameters=auto_fit_parameters,
                                                                         verbose=verbose if self.ds_aux.n_jobs == 1 else False),
                                                                 [self.feature_woes[feature] for feature in features]), total=len(features))):
                            self.feature_woes[features[i]].is_active, check_dfs[num_p][features[i]] = result
                        gc.collect()
                else:
                    for feature in tqdm(features):
                        self.feature_woes[feature].is_active, check_dfs[num_p][feature] = \
                            self.feature_woes[feature].auto_fit(verbose=verbose, **auto_fit_parameters)
    
                f_gini[num_p], Vlogic_features[num_p] = save_check_dfs(dfs=check_dfs[num_p], woe_best_samples=woe_best_samples,
                                                                       check_file=f'{folder}checks{f"_{num_p}" if len(params_list) > 1 else ""}.xlsx')
                woe_dfs = [check_dfs[num_p][f][5] for f in check_dfs[num_p] if self.feature_woes[f].is_active]
                if woe_dfs:
                    scorecard = pd.concat(woe_dfs)
                    scorecard.to_excel(f'{folder}scorecard{f"_{num_p}" if len(params_list) > 1 else ""}_all.xlsx', index=False)
                    scorecard = scorecard[scorecard.groupby(['feature'])['iteration'].transform('max') == scorecard['iteration']]
                    scorecard.to_excel(f'{folder}scorecard{f"_{num_p}" if len(params_list) > 1 else ""}.xlsx', index=False)
                    #scorecard[f'Gini {woe_best_samples[0] if len(woe_best_samples) == 1 else f"avg {woe_best_samples}"}'] = scorecard['feature'].map(f_gini[1]).round(self.round_woe)
                else:
                    print(f'There are no successfully binned features with parameters set {num_p}!')
    
            if len(params_list) > 1:
                WOE_best_dfs = []
                for f in features:
                    ginis = {num_p: f_gini[num_p][f] if f not in Vlogic_features[num_p] else f_gini[num_p][f] - BL_allow_Vlogic_to_increase_gini
                             for num_p in f_gini if f in f_gini[num_p]}
                    num_p = max(ginis, key=ginis.get)
                    tmp = check_dfs[num_p][f][5]
                    tmp['params_set'] = num_p
                    tmp[f'Gini {woe_best_samples[0] if len(woe_best_samples) == 1 else f"avg {woe_best_samples}"}'] = round(f_gini[num_p][f], self.round_woe)
                    WOE_best_dfs.append(tmp[tmp.groupby(['feature'])['iteration'].transform('max') == tmp['iteration']])
                if WOE_best_dfs:
                    scorecard = pd.concat(WOE_best_dfs)
                else:
                    print(f'There are no successfully binned features with any parameters set!')
            if not scorecard.empty:
                scorecard.to_excel(file_scorecard, index=False)
                self.import_scorecard(scorecard, verbose=False, fit_flag=False)
            excluded = [f for f in features if not self.feature_woes[f].is_active]
            if excluded:
                print(f'Excluded features {excluded} because no suitable binning was found for them')
        # --------------------------------------- autobinning for cross features ---------------------------------------
        if cross_features or cross_num_second_level != 0:
            print(f'Performing autobinning for cross features with parameters space of size {len(params_list)}...')
            check_dfs = {}
            f_gini = {}
            features_active = set([f for f in self.feature_woes if self.feature_woes[f].is_active])
            if scorecard.empty:
                scorecard = self.export_scorecard()
            new_sc = pd.DataFrame()
            if cross_features:
                crosses = {}
                for f in cross_features:
                    f1, f2 = cross_split(f)
                    if f1 in features_active:
                        try:
                            crosses[f1].append(f2)
                        except:
                            crosses[f1] = [f2]
                    else:
                        print(f'No active binning for feature {f1} found. Cross feature {f} skipped.')
            else:
                if cross_features_first_level is None:
                    cross_features_first_level = features if features else features_active
                cross_features_first_level = set(cross_features_first_level) & features_active
                if cross_num_second_level == -1 or cross_num_second_level >= len(features_active) - 1:
                    cross_num_second_level = len(features_active) - 1
                    crosses = {f1: [f2 for f2 in features_active if f2 != f1] for f1 in cross_features_first_level}
                else:
                    crosses = {}
                    print('Finding the best pairs to first-level features...')
                    for f1 in tqdm(cross_features_first_level):
                        df1_group = self.feature_woes[f1].ds.samples[self.feature_woes[f1].ds.train_name][['group']]
                        impurity = {}
                        if self.ds_aux.n_jobs > 1 and len(features) > self.ds_aux.n_jobs:
                            f2_features = [f2 for f2 in features_active if f2 != f1]
                            with futures.ProcessPoolExecutor(max_workers=self.ds_aux.n_jobs) as pool:
                                for i, result in enumerate(pool.map(partial(self.get_f2_impurity, df1_group=df1_group),
                                                 [self.feature_woes[f2] for f2 in f2_features])):
                                    impurity[f2_features[i]] = result
                                gc.collect()
                        else:
                            impurity = {f2: self.get_f2_impurity(self.feature_woes[f2], df1_group) for f2 in features_active if f2 != f1}
                        crosses[f1] = sorted(impurity, key=impurity.get, reverse=True)[:cross_num_second_level]
                cross_features = [cross_name(f1, f2) for f1 in crosses for f2 in crosses[f1]]
            auto_fit_parameters['BL_allow_Vlogic_to_increase_gini'] = 100
            auto_fit_parameters['max_n_bins'] = max(auto_fit_parameters['max_n_bins'], 5)
            check_dfs = {num_p: {} for num_p in range(1, len(params_list) + 1)}
            checks_result = {num_p: {} for num_p in range(1, len(params_list) + 1)}
            print('Creating feature_crosses...')
            for f1 in tqdm(crosses):
                self.feature_crosses[f1] = FeatureCross(self.feature_woes[f1])
                self.feature_crosses[f1].add_f2_features([self.feature_woes[f2] for f2 in crosses[f1]])

            for num_p, param in enumerate(params_list, start=1):
                auto_fit_parameters.update({params_names[k]: param[k] for k in range(len(param))})
                print(f'Using parameters set {num_p}/{len(params_list)}: {auto_fit_parameters}')
                for p in ['SM_target_threshold', 'SM_size_threshold']:
                    if auto_fit_parameters[p] and auto_fit_parameters[p] < 1:
                        auto_fit_parameters[p] = int(round(len(self.ds_aux.samples[self.ds_aux.train_name]) * auto_fit_parameters[p], 0))
                print(f'Processing {len(crosses)} first level features on {self.ds_aux.n_jobs} CPU{"s" if self.ds_aux.n_jobs > 1 else ""}...')

                if self.ds_aux.n_jobs > 1 and len(features) > self.ds_aux.n_jobs:
                    with futures.ProcessPoolExecutor(max_workers=self.ds_aux.n_jobs) as pool:
                        crosses_list = list(crosses)
                        for i, result in enumerate(tqdm(pool.map(partial(self.auto_fit_feature, auto_fit_parameters=auto_fit_parameters,
                                                                         verbose=verbose if self.ds_aux.n_jobs == 1 else False),
                                                                 [self.feature_crosses[f1] for f1 in crosses_list]), total=len(crosses_list))):
                            checks_result[num_p][crosses_list[i]], check_dfs[num_p][crosses_list[i]] = result
                        gc.collect()
                else:
                    for f1 in tqdm(crosses):
                        checks_result[num_p][f1], check_dfs[num_p][f1] = self.feature_crosses[f1].auto_fit(verbose=verbose, **auto_fit_parameters)

                for f1 in crosses:
                    if check_dfs[num_p][f1]:
                        for i in range(5):
                            check_dfs[num_p][f1][i].insert(loc=0, column='feature1', value=f1)
                f_gini[num_p], _ = save_check_dfs(dfs=check_dfs[num_p], woe_best_samples=woe_best_samples, check_file=f'{folder}checks_crosses{f"_{num_p}" if len(params_list) > 1 else ""}.xlsx')
                woe_dfs = [check_dfs[num_p][f][5] for f in check_dfs[num_p] if check_dfs[num_p][f]]
                if woe_dfs:
                    new_sc = pd.concat(woe_dfs)
                    new_sc = new_sc[~new_sc['feature'].isin([cross_name(f1, f2) for f1 in checks_result[num_p] for f2 in checks_result[num_p][f1] if not checks_result[num_p][f1][f2]])]
                    new_sc.to_excel(f'{folder}scorecard_crosses{f"_{num_p}" if len(params_list) > 1 else ""}.xlsx', index=False)
                else:
                    print(f'There are no successfully binned cross features with parameters set {num_p}!')
            if len(params_list) > 1:
                woe_dfs = []
                others_dfs = {}
                for k in {k for i in f_gini for k in f_gini[i]}:
                    g = {i: f_gini[i][k] for i in f_gini if k in f_gini[i] and checks_result[i][k[0]][k[2]]}
                    if not g:
                        continue
                    num_p = max(g, key=g.get)
                    tmp = check_dfs[num_p][k[0]][5]
                    tmp['params_set'] = num_p
                    others_dfs[k[0]] = tmp[tmp['group'] == 'others'].copy()
                    tmp = tmp[(tmp['values'].str.split(' & ').str[0] == k[1]) & (tmp['feature'].str[6:].str.split('&').str[1] == k[2])]
                    woe_dfs.append(tmp)
                woe_dfs += list(others_dfs.values())
                if woe_dfs:
                    new_sc = pd.concat(woe_dfs).sort_values(by=['feature', 'group'])
                else:
                    print(f'There are no successfully binned cross features with any parameters set!')
            self.feature_crosses.clear()
            if not new_sc.empty:
                scorecard = pd.concat([scorecard, new_sc])
                scorecard.to_excel(file_scorecard, index=False)
                self.import_scorecard(scorecard, verbose=False, fit_flag=False)
        print(f'Scorecard saved to the file {file_scorecard}')
        if plot_flag != -1:
            print('Plotting binnings...')
            self.plot_bins(folder=folder + 'Figs_binning', features=features+cross_features, plot_flag=plot_flag, verbose=True)
        print(f'All done!{f" {len(features) - len(excluded)}/{len(features)} features successfully binned." if features else ""}'
              f'{f" Found {len([f for f in self.cross_features if f in cross_features])} cross features." if cross_features else ""}')

    @staticmethod
    def plot_bins_feature(feature_woe, ds_aux, folder=None, plot_flag=True, show_groups=False, all_samples=False):
        """
        Отрисовывает биннингодной переменной
        :param feature_woe: объект класса FeatureWOE
        :param ds_aux: вспомогательный ДатаСэмпл с полем среза
        :param folder: папка, в которую должны быть сохранены рисунки. По умолчанию не сохраняются
        :param plot_flag: флаг для вывода рисунка
        :param show_groups: флаг для отображения номер групп на рисунке
        """
        f_WOE = add_suffix(feature_woe.feature)
        ds = feature_woe.transform()
        if ds.samples is None:
            return None
        if not feature_woe.categorical_type:
            if ds_aux.time_column is None:
                fig, (ax_2, ax_1) = plt.subplots(1, 2, gridspec_kw={'width_ratios': [1, 3.5]}, figsize=(13.5, 6))
            else:
                fig, axd = plt.subplot_mosaic([['ul', 'ur'], ['l', 'l']], figsize=(13.5, 9),
                                              gridspec_kw={'width_ratios': [1, 3.5], 'height_ratios': [2, 1]})
                ax_1 = axd['ur']
                ax_2 = axd['ul']
                ax_3 = axd['l']
            x = ds.samples[ds.train_name][feature_woe.feature]
            if feature_woe.special_bins:
                x = x.replace(list(feature_woe.special_bins.values()), np.nan).dropna()
            x_stat = [x.min(), x.median(), x.max()]
            xlim = [x.quantile(0.05), x.quantile(0.95)]
            if x.min() >= 0:
                xlim[0] = x.min()
            x = x[(xlim[0] <= x) & (x <= xlim[1])]
            x.plot.hist(ax=ax_2, bins=20, rwidth=0.7, xlim=xlim)
            try:
                x.plot(kind='kde', ax=ax_2, secondary_y=True, xlim=xlim, color='green')
            except:
                pass
            ax_2.set_ylabel('Observations amount')
            ax_2.grid(True)
            ax_2.right_ax.set_yticklabels([])
            ax_2.right_ax.set_ylim(0, ax_2.right_ax.get_ylim()[1])
            ax_2.right_ax.grid(False)
            for tick in ax_2.get_xticklabels():
                tick.set_rotation(30)
            ax_2.tick_params(axis='both', which='both', length=5, labelbottom=True)
            ax_2.annotate(f'Min: {round(x_stat[0], feature_woe.round_digits)}', xy=(0, 1), xycoords=('axes fraction', 'axes fraction'),
                          xytext=(0, 59), textcoords='offset pixels', color='black', ha='left')
            ax_2.annotate(f'Median: {round(x_stat[1], feature_woe.round_digits)}', xy=(0, 1),
                          xycoords=('axes fraction', 'axes fraction'),
                          xytext=(0, 37), textcoords='offset pixels', color='black', ha='left')
            ax_2.annotate(f'Max: {round(x_stat[2], feature_woe.round_digits)}', xy=(0, 1),
                          xycoords=('axes fraction', 'axes fraction'),
                          xytext=(0, 15), textcoords='offset pixels', color='black', ha='left')
        else:
            if ds_aux.time_column is None:
                fig, ax_1 = plt.subplots(1, 1, figsize=(13.5, 6))
            else:
                fig, (ax_1, ax_3) = plt.subplots(2, 1, gridspec_kw={'height_ratios': [2, 1]}, figsize=(13.5, 9))
        if ds_aux.time_column is not None:
            gini_in_time = ds.calc_gini_in_time(ds_aux=ds_aux)
            gini_in_time.columns = [x[1] for x in gini_in_time.columns]
            gini_in_time[[c for c in gini_in_time.columns if c != 'Bootstrap std']].plot(ax=ax_3, marker='o')
            if 'Bootstrap std' in gini_in_time.columns:
                ax_3.fill_between(gini_in_time.index,
                                  gini_in_time['Bootstrap mean'] - 1.96 * gini_in_time['Bootstrap std'],
                                  gini_in_time['Bootstrap mean'] + 1.96 * gini_in_time['Bootstrap std'],
                                  alpha=0.1, color='blue', label='95% conf interval')
            ax_3.set_ylabel('Gini')
            ax_3.set_title('Gini Stability', fontsize=14)
            for tick in ax_3.get_xticklabels():
                tick.set_rotation(30)
            ax_3.tick_params(axis='both', which='both', length=5, labelbottom=True)
            ax_3.xaxis.get_label().set_visible(False)
            ax_3.legend(loc='center left', bbox_to_anchor=(1, 0.5))
        if ds_aux.id_column is not None:
            to_calc = feature_woe.ds.samples[ds.train_name].copy()
            to_calc[ds_aux.id_column] = ds_aux.samples[ds_aux.train_name][ds_aux.id_column]
            to_calc[f'{ds_aux.id_column}_1'] = to_calc.apply(lambda row: row[ds_aux.id_column] if row[ds.target] == 1 else np.nan, axis=1)
            stats = to_calc.groupby('group').agg({ds.target: ['sum', 'size'],
                                                  ds_aux.id_column: 'nunique',
                                                  f'{ds_aux.id_column}_1': 'nunique'}).set_axis(['target_0', 'Observations_0', 'Unique ids', 'Class 1 unique ids'], axis=1)
        else:
            stats = feature_woe.ds.samples[ds.train_name].groupby('group').agg({ds.target: ['sum', 'size']}).set_axis(['target_0', 'Observations_0'], axis=1)
            stats['Unique ids'] = stats['Observations_0']
            stats['Class 1 unique ids'] = stats['target_0']
        stats = stats.reset_index()
        samples_to_show = [ds.train_name]
        if all_samples:
            samples_to_show += [s for s in feature_woe.ds.samples if s != ds.train_name]
            for i, sample in enumerate(samples_to_show[1:], start=1):
                tmp = feature_woe.ds.samples[sample].copy()
                tmp['group'] = feature_woe.set_groups(data=tmp[feature_woe.feature])
                stats = stats.merge(tmp.groupby('group').agg({ds.target: ['sum', 'size']}).set_axis([f'target_{i}', f'Observations_{i}'], axis=1).reset_index(), on='group', how='outer')
                if all_samples > 1 and i > 0:
                    stats[f'woe_{i}'] = stats['group'].map(feature_woe.calc_simple_woes(df=tmp))
        stats['woe_0'] = stats['group'].map(feature_woe.woes)
        stats['down_trend'] = ((stats['woe_0'] < stats['woe_0'].shift(1)) & (stats['group'].shift(1) >= 0)) if not feature_woe.categorical_type else False
        stats['label'] = stats['group'].map(feature_woe.groups).astype('str').apply(lambda x: '\n'.join(wrap(str(x), 25)))
        stats['Target rate'] = (stats['target_0']*100 / stats['Observations_0']).round(2).astype('str') + '%'
        stats = stats.fillna(0)
        if feature_woe.categorical_type:
            stats = stats[pd.isnull(stats['woe_0']) == False].sort_values('woe_0').reset_index()
        if feature_woe.missing_group != -1 and not feature_woe.categorical_type:
            stats.loc[stats.group == feature_woe.missing_group, 'label'] += '\nwith missings'
        if not feature_woe.categorical_type:
            stats['label'] = stats['label'].replace({']': ')', '\[-inf': '(-inf'}, regex=True)
        if show_groups:
            stats['label'] += '\n' + 'group ' + stats['group'].astype('str')
        for i, group in enumerate(feature_woe.special_bins.keys()):
            stats.loc[stats['group'] == -2 - i, 'label'] = group
        ax_1.set_ylabel('Observations')
        ax_1.set_xticks(range(stats.shape[0]))
        ax_1.set_xticklabels(stats['label'], rotation=30, ha='right', fontsize=10 if show_groups else 11)
        ax2 = ax_1.twinx()
        ax2.set_ylabel('WOE', loc='bottom')
        ax2.grid(axis='y', zorder=1, alpha=0.6)
        for i, sample in enumerate(samples_to_show):
            w = 0.8/len(samples_to_show)
            shift = (i - (len(samples_to_show) - 1)/2)*(w + 0.03)
            amt = stats[f'Observations_{i}'].sum()
            ax_1.bar(stats.index + shift, (stats[f'Observations_{i}'] - stats[f'target_{i}']) / amt, width=w, zorder=0, alpha=1 - i*0.25, color='forestgreen', label=f'Class 0{" (" + sample + ")" if len(samples_to_show) > 1 else ""}')
            ax_1.bar(stats.index + shift, stats[f'target_{i}'] / amt, bottom=(stats[f'Observations_{i}'] - stats[f'target_{i}']) / amt, width=w, zorder=0, alpha=1 - i*0.25, color='indianred', label=f'Class 1{" (" + sample + ")" if len(samples_to_show) > 1 else ""}')
            if all_samples > 1 or i == 0:
                if all_samples <= 1:
                    shift = 0
                for x, y in stats[f'woe_{i}'].items():
                    ax2.annotate(str(y), xy=(x + shift, y), xytext=(30 if stats['down_trend'][x] else -30, 10), textcoords='offset pixels',
                                 color='black', ha='center', size=11 if (all_samples <= 1 or len(samples_to_show) < 3) else 10)
                if i == 0:
                    if feature_woe.categorical_type:
                        ax2.plot(stats.index + shift, stats['woe_0'], 'bo', linewidth=2.0, zorder=4, label='WOE')
                    else:
                        stats['line'] = stats['group'].apply(lambda x: not isinstance(x, str) and x >= 0).astype('int')
                        ax2.plot(stats[stats['line'] == 1].index + shift, stats[stats['line'] == 1]['woe_0'], 'bo-', linewidth=2.0, zorder=4, label='WOE')
                        ax2.plot(stats[stats['line'] == 0].index + shift, stats[stats['line'] == 0]['woe_0'], 'bo', linewidth=2.0, zorder=4)
                else:
                    ax2.plot(stats.index + shift, stats[f'woe_{i}'], 'bo', linewidth=2.0, zorder=4, alpha=1 - i*0.25)
            if i == 0 and feature_woe.ds.bootstrap_base is not None:
                feature_woe.ds.bootstrap_base['group'] = feature_woe.set_groups(data=feature_woe.ds.bootstrap_base[feature_woe.feature])
                bootstrap_woes_l = [feature_woe.calc_simple_woes(df=feature_woe.ds.bootstrap_base.iloc[idx]) for idx in feature_woe.ds.bootstrap]
                bootstrap_woes = {g: np.array([w[g] for w in bootstrap_woes_l if g in w]) for g in feature_woe.groups}
                stats['woe_mean'] = stats['group'].map({g: np.mean(v) for g, v in bootstrap_woes.items()})
                stats['woe_std'] = stats['group'].map({g: np.std(v) for g, v in bootstrap_woes.items()})
                ax2.errorbar(stats.index + shift, stats['woe_mean'], xerr=0, yerr=stats['woe_std'] * 1.96, color='b', capsize=5, fmt='none', alpha=0.4)

        for stat, shift in {'Target rate': 15, 'Class 1 unique ids': 37, 'Unique ids': 59, 'Observations_0': 83}.items():
            ax_1.annotate(stat.replace('_0', ''), xy=(-0.5, 1), xycoords=('data', 'axes fraction'),
                          xytext=(0, shift), textcoords='offset pixels', color='black', ha='right')
            for i, val in enumerate(stats[stat].values):
                ax_1.annotate(str(val), xy=(i, 1), xycoords=('data', 'axes fraction'),
                              xytext=(0, shift), textcoords='offset pixels', color='black', ha='center')
        ax_1.annotate('Gini: %s' % ('; '.join(['%s %.2f' % (name, gini) for name, gini in ds.calc_gini(features=[f_WOE]).T[f_WOE].to_dict().items()])),
                      xy=(0.5, 1), xycoords=('figure fraction', 'axes fraction'), xytext=(0, 115), textcoords='offset pixels', color='blue', ha='center')
        ax_1.grid(False)
        ax_1.tick_params(axis='y', which='both', length=5)
        h1, l1 = ax_1.get_legend_handles_labels()
        h2, l2 = ax2.get_legend_handles_labels()
        ax_1.legend(h1 + h2, l1 + l2, bbox_to_anchor=(1.27 if len(samples_to_show) > 1 else 1.23, 1), fontsize=9 if len(samples_to_show) > 1 else 11)
        plt.suptitle(ds_aux.feature_titles[feature_woe.feature] if feature_woe.feature in ds_aux.feature_titles
                     else feature_woe.feature, fontsize=13, weight='bold')
        fig.tight_layout()
        if folder is not None:
            fig.savefig(f'{folder}/{f_WOE}.png', bbox_inches="tight")
        if plot_flag:
            plt.show()
        return fig

    @staticmethod
    def plot_bins_feature_cross(params, ds_aux, folder=None, plot_flag=True, show_groups=False,
                                all_samples=False):
        """
        Отрисовывает биннингодной переменной
        :param feature_woe: объект класса FeatureWOE
        :param ds_aux: вспомогательный ДатаСэмпл с полем среза
        :param folder: папка, в которую должны быть сохранены рисунки. По умолчанию не сохраняются
        :param plot_flag: флаг для вывода рисунка
        :param show_groups: флаг для отображения номер групп на рисунке
        """
        feature_cross, f2 = params
        feature_woe = feature_cross.feature_woe
        feature_woe2 = list(feature_cross.cross[f2].values())[0]
        f = cross_name(feature_woe.feature, f2)
        f_WOE = add_suffix(f)
        ds = feature_cross.transform(features=[f2])
        size = 12 - len(feature_cross.get_woes(f2)) // 7
        group_coord = {}
        group_w = {}
        group_d = {}
        for i, (g, fw) in enumerate(feature_cross.cross[f2].items()):
            for j, g2 in enumerate(fw.groups):
                w = 0.8 / len(fw.groups)
                shift = (j - (len(fw.groups) - 1) / 2) * (w + 0.03)
                group_coord[str([g, g2])] = i + shift
                group_w[str([g, g2])] = w
                if len(fw.groups) == 1:
                    group_d[g] = [i - w / 2, i + w / 2]
                elif j == 0:
                    group_d[g] = [i + shift - w / 2, ]
                elif j == len(fw.groups) - 1:
                    group_d[g].append(i + shift + w / 2)
        if ds.samples is None:
            return None

        if ds_aux.time_column is None:
            fig, (ax_1, ax_x) = plt.subplots(2, 1, gridspec_kw={'height_ratios': [6, 0.5]}, figsize=(13.5, 6.5))
        else:
            fig, (ax_1, ax_x, ax_3) = plt.subplots(3, 1, gridspec_kw={'height_ratios': [6, 0.5, 3]},
                                                   figsize=(13.5, 9.5))
        if ds_aux.time_column is not None:
            gini_in_time = ds.calc_gini_in_time(ds_aux=ds_aux)
            gini_in_time.columns = [x[1] for x in gini_in_time.columns]
            gini_in_time[[c for c in gini_in_time.columns if c != 'Bootstrap std']].plot(ax=ax_3, marker='o')
            if 'Bootstrap std' in gini_in_time.columns:
                ax_3.fill_between(gini_in_time.index,
                                  gini_in_time['Bootstrap mean'] - 1.96 * gini_in_time['Bootstrap std'],
                                  gini_in_time['Bootstrap mean'] + 1.96 * gini_in_time['Bootstrap std'],
                                  alpha=0.1, color='blue', label='95% conf interval')
            ax_3.set_ylabel('Gini')
            ax_3.set_title('Gini Stability', fontsize=14)
            for tick in ax_3.get_xticklabels():
                tick.set_rotation(30)
            ax_3.tick_params(axis='both', which='both', length=5, labelbottom=True)
            ax_3.xaxis.get_label().set_visible(False)
            ax_3.legend(loc='center left', bbox_to_anchor=(1, 0.5))
        to_calc = feature_woe.ds.samples[ds.train_name].merge(
            pd.concat([fw.ds.samples[ds.train_name][f2] for g, fw in feature_cross.cross[f2].items()]), left_index=True,
            right_index=True)
        to_calc['group'] = feature_cross.set_groups(data=to_calc, f2=f2)
        if ds_aux.id_column is not None:
            to_calc[ds_aux.id_column] = ds_aux.samples[ds_aux.train_name][ds_aux.id_column]
            to_calc[f'{ds_aux.id_column}_1'] = to_calc.apply(
                lambda row: row[ds_aux.id_column] if row[ds.target] == 1 else np.nan, axis=1)
            stats = to_calc.groupby('group').agg({ds.target: ['sum', 'size'],
                                                  ds_aux.id_column: 'nunique',
                                                  f'{ds_aux.id_column}_1': 'nunique'}).set_axis(
                ['target_0', 'Observations_0', 'Unique ids', 'Class 1 unique ids'], axis=1)
        else:
            stats = to_calc.groupby('group').agg({ds.target: ['sum', 'size']}).set_axis(['target_0', 'Observations_0'],
                                                                                        axis=1)
            stats['Unique ids'] = stats['Observations_0']
            stats['Class 1 unique ids'] = stats['target_0']
        stats = stats.reset_index()
        samples_to_show = [ds.train_name]
        if all_samples:
            samples_to_show += [s for s in feature_woe.ds.samples if s != ds.train_name]
            for i, sample in enumerate(samples_to_show[1:], start=1):
                tmp = feature_woe.ds.samples[sample].merge(
                    pd.concat([fw.ds.samples[sample][f2] for g, fw in feature_cross.cross[f2].items()]),
                    left_index=True, right_index=True)
                tmp['group'] = feature_cross.set_groups(data=tmp, f2=f2)
                stats = stats.merge(tmp.groupby('group').agg({ds.target: ['sum', 'size']}).set_axis(
                    [f'target_{i}', f'Observations_{i}'], axis=1).reset_index(), on='group', how='outer')
                if all_samples > 1 and i > 0:
                    stats[f'woe_{i}'] = stats['group'].map(feature_woe.calc_simple_woes(df=tmp))
        stats['woe_0'] = stats['group'].map(feature_cross.get_woes(f2))
        stats['label'] = stats['group'].map(feature_cross.get_values(f2)).astype('str').apply(
            lambda x: '\n'.join(wrap(str(x), 25)))
        stats['Target rate'] = (stats['target_0'] * 100 / stats['Observations_0']).round(1).astype('str') + '%'
        stats['coord'] = stats['group'].map(group_coord)
        stats['w'] = stats['group'].map(group_w)
        stats['group1'] = stats['group'].str[1:-1].str.split(', ').str[0].astype('int')
        stats['group2'] = stats['group'].str[1:-1].str.split(', ').str[1].astype('int')
        stats['down_trend'] = ((stats['woe_0'] < stats['woe_0'].shift(1)) & (
                    stats['group1'] == stats['group1'].shift(1)) & (stats['group2'].shift(1) >=0)) if not feature_woe2.categorical_type else False
        stats = stats.fillna(0)
        for g in feature_woe.groups:
            if feature_cross.cross[f2][g].missing_group != -1 and not feature_cross.cross[f2][g].categorical_type:
                stats.loc[(stats['group1'] == g) & (
                            stats['group2'] == feature_cross.cross[f2][g].missing_group), 'label'] += '\nwith missings'
        if not feature_woe2.categorical_type:
            stats['label'] = stats['label'].replace({']': ')', '\[-inf': '(-inf'}, regex=True)
        if show_groups:
            stats['label'] += '\n' + 'group ' + stats['group'].astype('str')
        for i, group in enumerate(feature_woe.special_bins.keys()):
            stats.loc[stats['group'] == -2 - i, 'label'] = group
        ax_1.set_ylabel('Observations')
        ax_1.set_xticks(stats['coord'])
        ax_1.set_xticklabels(stats['label'], rotation=30, ha='right', fontsize=size - 1 if show_groups else size)

        ax2 = ax_1.twinx()
        ax2.set_ylabel('WOE', loc='bottom')
        ax2.grid(axis='y', zorder=1, alpha=0.6)
        for i, sample in enumerate(samples_to_show):
            w = stats['w'] * 0.9 / len(samples_to_show)
            shift = (i - (len(samples_to_show) - 1) / 2) * (w + 0.03)
            amt = stats[f'Observations_{i}'].sum()
            ax_1.bar(stats['coord'] + shift, (stats[f'Observations_{i}'] - stats[f'target_{i}']) / amt, width=w,
                     zorder=0, alpha=1 - i * 0.25, color='forestgreen',
                     label=f'Class 0{" (" + sample + ")" if len(samples_to_show) > 1 else ""}')
            ax_1.bar(stats['coord'] + shift, stats[f'target_{i}'] / amt,
                     bottom=(stats[f'Observations_{i}'] - stats[f'target_{i}']) / amt, width=w, zorder=0,
                     alpha=1 - i * 0.25, color='indianred',
                     label=f'Class 1{" (" + sample + ")" if len(samples_to_show) > 1 else ""}')
            if all_samples > 1 or i == 0:
                if all_samples <= 1:
                    shift = 0
                for x, y in stats[f'woe_{i}'].items():
                    ax2.annotate(str(y), xy=(stats['coord'][x] + shift, y),
                                 xytext=(30 if stats['down_trend'][x] else -30, 10), textcoords='offset pixels',
                                 color='black', ha='center',
                                 size=size if (all_samples <= 1 or len(samples_to_show) < 3) else size - 1)
                if i == 0:
                    if feature_woe2.categorical_type:
                        ax2.plot(stats['coord'] + shift, stats['woe_0'], 'bo', linewidth=2.0, zorder=4, label='WOE')
                    else:
                        for j, (g, group) in enumerate(stats.groupby('group1')):
                            group['line'] = group['group2'].apply(lambda x: not isinstance(x, str) and x >= 0).astype(
                                'int')
                            ax2.plot(group[group['line'] == 1]['coord'] + shift, group[group['line'] == 1]['woe_0'],
                                     'bo-', linewidth=2.0, zorder=4, label='WOE' if j == 0 else None)
                            ax2.plot(group[group['line'] == 0]['coord'] + shift, group[group['line'] == 0]['woe_0'],
                                     'bo', linewidth=2.0, zorder=4)
                else:
                    ax2.plot(stats['coord'] + shift, stats[f'woe_{i}'], 'bo', linewidth=2.0, zorder=4,
                             alpha=1 - i * 0.25)
            if i == 0 and feature_woe.ds.bootstrap_base is not None:
                tmp = feature_woe.ds.bootstrap_base.merge(feature_woe2.ds.bootstrap_base[f2], left_index=True,
                                                          right_index=True)
                tmp['group'] = feature_cross.set_groups(data=tmp, f2=f2)
                bootstrap_woes_l = [feature_woe.calc_simple_woes(df=tmp.iloc[idx]) for idx in feature_woe.ds.bootstrap]
                bootstrap_woes = {g: np.array([w[g] for w in bootstrap_woes_l if g in w]) for g in
                                  tmp['group'].unique()}
                stats['woe_mean'] = stats['group'].map({g: np.mean(v) for g, v in bootstrap_woes.items()})
                stats['woe_std'] = stats['group'].map({g: np.std(v) for g, v in bootstrap_woes.items()})
                ax2.errorbar(stats['coord'] + shift, stats['woe_mean'], xerr=0, yerr=stats['woe_std'] * 1.96, color='b',
                             capsize=5, fmt=',', alpha=0.4)

        for stat, shift in {'Target rate': 15, 'Class 1 unique ids': 37, 'Unique ids': 59,
                            'Observations_0': 83}.items():
            ax_1.annotate(stat.replace('_0', ''), xy=(-0.5, 1), xycoords=('data', 'axes fraction'),
                          xytext=(0, shift), textcoords='offset pixels', color='black', ha='right', size=size)
            for i, val in enumerate(stats[stat].values):
                ax_1.annotate(str(val), xy=(stats['coord'][i], 1), xycoords=('data', 'axes fraction'),
                              xytext=(0, shift), textcoords='offset pixels', color='black', ha='center', size=size)
        ax_1.annotate('Gini: %s' % ('; '.join(
            ['%s %.2f' % (name, gini) for name, gini in ds.calc_gini(features=[f_WOE]).T[f_WOE].to_dict().items()])),
                      xy=(0.5, 1), xycoords=('figure fraction', 'axes fraction'), xytext=(0, 115),
                      textcoords='offset pixels', color='blue', ha='center')
        ax_1.grid(False)
        ax_1.tick_params(axis='y', which='both', length=5)
        h1, l1 = ax_1.get_legend_handles_labels()
        h2, l2 = ax2.get_legend_handles_labels()
        ax_1.legend(h1 + h2, l1 + l2, bbox_to_anchor=(1.27 if len(samples_to_show) > 1 else 1.15, 1),
                    fontsize=9 if len(samples_to_show) > 1 else 11)
        ds_aux.add_cross_description([f])
        plt.suptitle(ds_aux.feature_titles[f] if f in ds_aux.feature_titles else f, fontsize=13, weight='bold')
        fig.tight_layout()
        ax_x.set_xlim(ax_1.get_xlim())
        ax_x.set_yticks([])
        ax_x.axis('off')
        ax_1.annotate(f2, xy=(1, 0), xytext=(10, -30), textcoords='offset pixels', ha='left', va='top',
                      xycoords='axes fraction')
        ax_x.annotate(feature_woe.feature, xy=(1, 1), xytext=(10, 10), xycoords='axes fraction',
                      textcoords='offset pixels', ha='left', va='top')
        for g, d in group_d.items():
            ax_x.annotate('', xy=(d[0], 1), xytext=(d[1], 1), xycoords=('data', 'axes fraction'),
                          arrowprops={'arrowstyle': '|-|', 'facecolor': 'red'}, annotation_clip=False)
            label = '\n'.join(wrap(str(feature_woe.groups[g]), 100 // len(feature_woe.groups)))
            if not feature_woe.categorical_type:
                label = label.replace(']', ')').replace('[-inf', '(-inf')
            if feature_woe.missing_group != -1 and g == feature_woe.missing_group:
                label += '\nwith missings'
            ax_x.annotate(label, xy=((d[0] + d[1]) / 2, 1), xytext=(0, -10), xycoords=('data', 'axes fraction'),
                          textcoords='offset pixels', ha='center', va='top',
                          size=size if len(label) < 100 else size - 1 if len(label) < 200 else size - 2)
        if folder is not None:
            fig.savefig(f'{folder}/{f_WOE}.png', bbox_inches="tight")
        if plot_flag:
            plt.show()
        return fig

    def plot_bins(self, features=None, folder=None, plot_flag=True, show_groups=False, verbose=False, all_samples=False):
        """
        Отрисовка биннинга
        :param features: список переменных для обработки. При None отрисоываются все активные переменные
        :param folder: папка, в которую должны быть сохранены рисунки. При None не сохраняются
        :param plot_flag: флаг для вывода рисунка
        :param show_groups: флаг для отображения номер групп на рисунке
        :param verbose: флаг для отображения счетчика обработанных рисунков
        :param all_samples: отрисовка бинов по всем сэмплам, может принимать значения:
                            0, False - строятся бины только на трэйне
                            1, True – строятся бины по всем сэмплам, таргет рейт указывается только для трэйна
                            >1  – строятся бины и таргет рейт указывается по всем сэмплам

        :return: список из графиков [plt.figure]
        """
        if folder and not os.path.exists(folder):
            os.makedirs(folder)
        if features is None:
            features = list(self.feature_woes) + list(self.cross_features)
        if plot_flag:
            verbose = False
        cross_features = [f for f in features if f in self.cross_features]
        features = [f for f in features if f in self.feature_woes and self.feature_woes[f].is_active]
        if self.ds_aux.n_jobs > 1 and len(features) > self.ds_aux.n_jobs:
            with futures.ProcessPoolExecutor(max_workers=self.ds_aux.n_jobs) as pool:
                pool_iter = pool.map(partial(self.plot_bins_feature, ds_aux=self.ds_aux, folder=folder,
                                             plot_flag=plot_flag, show_groups=show_groups, all_samples=all_samples),
                                     [self.feature_woes[f] for f in features])
                pool_iter_cross = pool.map(partial(self.plot_bins_feature_cross, ds_aux=self.ds_aux, folder=folder,
                                             plot_flag=plot_flag, show_groups=show_groups, all_samples=all_samples),
                                     [(self.feature_crosses[cross_split(f)[0]], cross_split(f)[1]) for f in cross_features])
                figs = list(tqdm(pool_iter, total=len(features)) if verbose else pool_iter) + list(tqdm(pool_iter_cross, total=len(cross_features)) if verbose else pool_iter_cross)
            gc.collect()
        else:
            figs = [self.plot_bins_feature(feature_woe=self.feature_woes[f], ds_aux=self.ds_aux, folder=folder,
                                           plot_flag=plot_flag, show_groups=show_groups, all_samples=all_samples)
                    for f in (tqdm(features) if verbose else features)] + \
                   [self.plot_bins_feature_cross(params=(self.feature_crosses[cross_split(f)[0]], cross_split(f)[1]),
                                                 ds_aux=self.ds_aux, folder=folder, plot_flag=plot_flag, show_groups=show_groups, all_samples=all_samples)
                    for f in (tqdm(cross_features) if verbose else cross_features)]
        plt.close('all')
        return figs

    def merge(self, feature, groups_list, plot_flag=True):
        """
        Объединение двух бинов
        :param feature: переменная
        :param groups_list: [group1, group2] - список из двух бинов для объединения
        :param plot_flag: флаг вывода графика после разделения
        """

        print(f'{feature}: merging groups {groups_list}')
        if feature in self.feature_woes:
            self.feature_woes[feature].merge(groups_list)
        elif feature in self.cross_features:
            f1, f2 = cross_split(feature)
            if isinstance(groups_list[0], list) and isinstance(groups_list[1], list):
                if groups_list[0][0] != groups_list[1][0]:
                    print(f'For cross features first level groups must match! {groups_list[0][0]} != {groups_list[1][0]}')
                    return None
                self.feature_crosses[f1].cross[f2][groups_list[0][0]].merge([groups_list[0][1], groups_list[1][1]])
            else:
                print('For cross features groups_list must consist of two lists, for example [[0, 1], [0, 2]]')
                return None
        else:
            print(f'{feature} does not exist! Skipping...')
            return None
        if plot_flag:
            self.plot_bins(features=[feature], folder=None, plot_flag=True, show_groups=True)

    def merge_by_woe(self, woe_threshold=0.05):
        """
        Объединение всех близких по ВОЕ бинов
        :param woe_threshold: минимальная разрешенная дельта WOE между соседними бинами
        :param with_missing: должна ли выполняться проверка для бина с пустыми значениями
        """
        for feature in self.feature_woes:
            self.feature_woes[feature].merge_by_woe(woe_threshold=woe_threshold)

    def merge_by_size(self, target_threshold=5, size_threshold=100):
        """
        Объединение малых бинов с соседними ближайшими по ВОЕ
        :param target_threshold: минимальное кол-во (доля) наблюдений с целевым событием в бине
        :param size_threshold: минимальное кол-во (доля) наблюдений в бине
        """
        for feature in self.feature_woes:
            self.feature_woes[feature].merge_by_size(target_threshold=target_threshold, size_threshold=size_threshold)

    def split(self, feature, group=None, to_add=None, min_bin_size=0.05, criterion='entropy', scoring='neg_log_loss',
              plot_flag=True):
        """
        Разделение выбранного бина на две части
        :param feature: переменная
        :param group: номер бина для разделения, начиная с 0
        :param to_add: для числовых - граница между бинами, для категориальных - список значений для выделения в новый бин
        :param min_bin_size: минимальное число (доля) наблюдений в каждом бине
        :param criterion: критерий расщепления. Варианты значений: 'entropy', 'gini'
        :param scoring: метрика для оптимизации
        :param plot_flag: флаг вывода графика после разделения
        """

        print(f'{feature}: splitting group {group}')
        if feature in self.feature_woes:
            self.feature_woes[feature].split(group=group, to_add=to_add, min_bin_size=min_bin_size, criterion=criterion, scoring=scoring)
        elif feature in self.cross_features:
            f1, f2 = cross_split(feature)
            if isinstance(group, list):
                self.feature_crosses[f1].cross[f2][group[0]].split(group=group[1], to_add=to_add, min_bin_size=min_bin_size, criterion=criterion, scoring=scoring)
            else:
                print('For cross features group must be a list!')
                return None
        else:
            print(f'{feature} does not exist! Skipping...')
            return None
        if plot_flag:
            self.plot_bins(features=[feature], folder=None, plot_flag=True, show_groups=True)

    def show_history(self, feature):
        """
        Вывод истории биннинга одной переменной
        :param feature: название переменной
        """
        if feature not in self.feature_woes:
            print(f'{feature} not in included features!')
        else:
            curr_iteration = self.feature_woes[feature].curr_iteration()
            for i, h in enumerate(self.feature_woes[feature].history):
                print(f'Iteration {i}{" (current)" if i == curr_iteration else ""}')
                self.rollback(feature, iteration=i, plot_flag=True)
            if curr_iteration != len(self.feature_woes[feature].history) - 1:
                self.rollback(feature, iteration=curr_iteration, plot_flag=False)

    def rollback(self, feature, iteration=None, plot_flag=True):
        """
        Откат биннинга на одну из предыдущих итераций
        :param feature: пемеренная
        :param iteration: номер итерации. Возможны отрицательные значения: -1 - последняя итерация, -2 - предпоследняя и т.д.
        :param plot_flag: флаг вывода графика после отката
        """
        if feature not in self.feature_woes:
            print (f'{feature} does not exist! Skipping...')
        else:
            self.feature_woes[feature].rollback(iteration=iteration)
            if plot_flag:
                self.plot_bins(features=[feature], folder=None, plot_flag=True, show_groups=True)

    def transform(self, ds, features=None, verbose=False):
        """
        Трансформация ДатаСэмпла
        :param ds: ДатаСэмпл
        :param features: список переменных для трансформации. При None берутся ds.features для которых есть активный биннинг
        :param verbose: флаг для вывода комментариев в процессе работы

        :return: трансформированный ДатаСэмпл
        """
        cross_features = self.cross_features
        if features is None:
            features = {rem_suffix(f) for f in ds.features} | cross_features
        else:
            features = {rem_suffix(f) for f in features}
        new_features = {add_suffix(f) for f in features}
        if ds.samples is None or all([new_features.issubset(ds.samples[name].columns) for name in ds.samples]):
            ds.features = list(new_features)
            return ds
        if verbose > 0:
            print('Transforming features...')
        ds = copy.deepcopy(ds)
        for f in features:
            f_WOE = add_suffix(f)
            if f in self.feature_woes and self.feature_woes[f].is_active:
                for name in ds.samples:
                    ds.samples[name][f_WOE] = self.feature_woes[f].set_avg_woes(data=ds.samples[name][f])
                if ds.bootstrap_base is not None:
                    ds.bootstrap_base[f_WOE] = self.feature_woes[f].set_avg_woes(data=ds.bootstrap_base[f])
            elif f in cross_features:
                f1, f2 = cross_split(f)
                for name in ds.samples:
                    ds.samples[name][f_WOE] = self.feature_crosses[f1].set_avg_woes(data=ds.samples[name][[f1, f2]], f2=f2)
                if ds.bootstrap_base is not None:
                    ds.bootstrap_base[f_WOE] = self.feature_crosses[f1].set_avg_woes(data=ds.bootstrap_base[[f1, f2]], f2=f2)
            else:
                if verbose > 0:
                    print(f"Can't transform feature {f}. Skipped.")
                new_features.remove(f_WOE)
                continue
        ds.features = list(new_features)
        return ds

    def export_scorecard(self, out=None, features=None, full=True, history=False):
        """
        Сохранение биннинга в файл
        :param out: название файла
        :param features: список переменных для сохранения. При None сохраняются все, имеющие активный биннинг
        :param full: если True, то добавляет в файл поля с доп статистикой по бинам
        :param history: если True, то сохраняется вся история биннингов. Дубли биннингов удаляются, текущий биннинг записывается последней итерацией
        :return: датафрейм со скоркартой
        """
        cross_features = self.cross_features
        if features is None:
            features = [f for f in self.feature_woes if self.feature_woes[f].is_active] + list(cross_features)
        dfs = []
        for f in features:
            if f in self.feature_woes and self.feature_woes[f].is_active:
                if history:
                    curr_iteration = self.feature_woes[f].curr_iteration()
                    processed = []
                    i2 = 0
                    for i, h in enumerate(self.feature_woes[f].history):
                        if h not in processed and h != self.feature_woes[f].history[curr_iteration]:
                            self.feature_woes[f].rollback(iteration=i)
                            dfs.append(self.feature_woes[f].export_scorecard(full=full, iteration=i2))
                            processed.append(h)
                            i2 += 1
                    self.feature_woes[f].rollback(iteration=curr_iteration)
                    dfs.append(self.feature_woes[f].export_scorecard(full=full, iteration=i2))
                else:
                    dfs.append(self.feature_woes[f].export_scorecard(full=full))
            elif f in cross_features:
                f1, f2 = cross_split(f)
                dfs.append(self.feature_crosses[f1].export_scorecard(features=[f2], full=full))
            else:
                print(f'No active binning for feature {f} found. Skipped.')
        if dfs:
            df = pd.concat(dfs).reset_index(drop=True)
        else:
            df = pd.DataFrame()
        if out is not None:
            df.to_excel(out, index=False)
        return df

    def import_scorecard(self, scorecard, features=None, verbose=True, fit_flag=False):
        """
        Импорт биннинга из файла
        :param scorecard: путь к эксель файлу или датафрейм с готовыми биннингами для импорта
        :param features: список переменных для импорта биннинга. При None биннинг импортируется для всех, которые есть в файле
        :param verbose: флаг для вывода комментариев в процессе работы
        :param fit_flag: при True - WOE всех бинов перерасчитываются на текущей выборке
                         при False - WOE берутся из скоркарты. Если поле 'woe' в скоркарте отсутвует, то автоматически ставится fit_flag=True
        """
        if isinstance(scorecard, str):
            if scorecard[-5:] == '.xlsx' or scorecard[-4:] == '.xls':
                scorecard = pd.read_excel(scorecard)
            else:
                print('Unknown format of import file. Abort.')
                return None
        if features is None:
            features = list(scorecard['feature'].unique())
        for feature in features:
            if not is_cross(feature):
                if feature in self.feature_woes:
                    if verbose > 0:
                        print(f'Replacing binning for {feature}..')
                    self.feature_woes[feature].import_scorecard(scorecard[scorecard['feature'] == feature], verbose=verbose, fit_flag=fit_flag)
                else:
                    print(f'Feature {feature} is missing in self.feature_woes! Skipped.')
        scorecard = scorecard[(scorecard['feature'].str.startswith('cross_')) & (scorecard['feature'].str.contains('&'))].copy()
        if not scorecard.empty:
            scorecard['f1'] = scorecard['feature'].str[6:].str.split('&').str[0]
            scorecard['f2'] = scorecard['feature'].str[6:].str.split('&').str[1]
            scorecard['categorical_type1'] = scorecard['categorical_type'].str[1:-1].str.split(', ').str[0].fillna('')
            scorecard['categorical_type2'] = scorecard['categorical_type'].str[1:-1].str.split(', ').str[1].fillna('')
            scorecard['group1'] = np.where(scorecard['group'] != 'others', scorecard['group'].str[1:-1].str.split(', ').str[0].fillna(0), 'others')
            scorecard['group2'] = np.where(scorecard['group'] != 'others', scorecard['group'].str[1:-1].str.split(', ').str[1].fillna(0), 'others')
            scorecard['values1'] = scorecard['values'].str.split(' & ').str[0]
            scorecard['values2'] = scorecard['values'].str.split(' & ').str[1]
            scorecard['missing1'] = np.where(scorecard['missing'].isin([0, 1, '0', '1']), 0,
                                             scorecard['missing'].astype('str').str[1:-1].str.split(', ').str[0].fillna('0')
                                             )
            scorecard['missing2'] = np.where(scorecard['missing'].isin([0, 1, '0', '1']), scorecard['missing'],
                                             scorecard['missing'].astype('str').str[1:-1].str.split(', ').str[1].fillna('0')
                                             )
            scorecard[['group1', 'group2', 'missing1', 'missing2']] = scorecard[['group1', 'group2', 'missing1', 'missing2']].applymap(lambda x: pd.to_numeric(x, errors='ignore'))
            for f1, scorecard_f1 in scorecard.groupby('f1'):
                if verbose > 0:
                    print(f'Replacing binning for cross features {f1}..')
                if f1 in self.feature_woes:
                    fw = copy.deepcopy(self.feature_woes[f1])
                else:
                    fw = self.create_feature_woe(DataSamples(), f1)
                fw.import_scorecard(scorecard=scorecard_f1[['f1', 'categorical_type1', 'group1', 'values1', 'woe', 'missing1']]\
                                    .set_axis(['feature', 'categorical_type', 'group', 'values', 'woe', 'missing'], axis=1)\
                                    .drop_duplicates(subset=['group']),
                                    verbose=False, fit_flag=False)
                self.feature_crosses[f1] = FeatureCross(fw)
                self.feature_crosses[f1].add_f2_features([self.feature_woes[f2] if f2 in self.feature_woes else self.create_feature_woe(DataSamples(), f2)
                                                             for f2 in scorecard_f1['f2'].unique()])
                self.feature_crosses[f1].import_scorecard(scorecard_f1, verbose=verbose, fit_flag=fit_flag)

    @property
    def cross_features(self):
        return {cross_name(f1, f2) for f1 in self.feature_crosses for f2 in self.feature_crosses[f1].cross}

#----------------------------------------------------------------------------------------------------------


class FeatureWOE:
    '''
    ВОЕ-трансформация для отдельного фактора. Для каждого фактора должен быть создан свой экземпляр
    '''
    def __init__(self, ds, feature, round_digits=3, round_woe=3, rounding_migration_coef=0.001,
                 simple=False, n_folds=5, woe_adjust=0.5, alpha=0, alpha_range=None, alpha_scoring='neg_log_loss',
                 alpha_best_criterion='min', missing_process='separate', missing_min_part=0.05,
                 others_process='missing', opposite_sign_to_others=False, special_bins=None):
        """
        :param ds: ДатаСэмпл, для которого будут рассчитываться биннинги
        :param feature: переменная

        ---Параметры для расчета WOE---
        :param round_digits: число знаков после запятой для округление границ бинов и WOE.
                 При округлении границ бинов происходит проверка на долю мигрирующих наблюдений. Если округление приедет к миграции большой доли наблюдений,
                 то round_digits увеличивается до тех пор, пока доля не упадет ниже rounding_migration_coef
        :param rounding_migration_coef: максимально допустимая доля наблюдений для миграции между бинами при округлении
        :param simple: если True, то расчет WOE происходит на трэйн сэмпле, иначе берется среднее значение по фолдам
        :param n_folds: кол-во фолдов для расчета WOE при simple=False
        :param woe_adjust: корректировочный параметр для расчета EventRate_i в бине i:
                           EventRate_i = (n1_i+woe_adjust)/(n0_i+woe_adjust),
                           где n1_i - кол-во наблюдений с target=1 в бине i, n0_i - кол-во наблюдений с target=0 в бине i
        :param alpha: коэффициент регуляризации для расчета WOE. Если задан, то WOE для бина i вычисляется по формуле:
                      SmoothedWOE_i = log((n + alpha)*EventRate/(n*EventRate_i + alpha)),
                      где n - число наблюдений, EventRate = n1/n0,
                      n1 - общее кол-во наблюдений с target=1, n0 - общее кол-во наблюдений с target=0
        :param alpha_range: если alpha=None, то подбирается оптимальное значение alpha из диапазона alpha_range
        :param alpha_scoring: метрика, используемая для оптимизации alpha
        :param alpha_best_criterion: 'min' - минимизация метрики alpha_scoring, 'max' - максимизация метрики

        ---Обработка пустых значений---
        :param missing_process: способ обработки пустых значений
                'separate' - помещать в отдельный бин
                'min' - объединять с бином с минимальным WOE
                'max' - объединять с бином с максимальным WOE
                'nearest' - объединять с ближайшим по WOE биномом
                'min_or_separate' - если доля пустых значений меньше missing_min_part, то объединять с бином с минимальным WOE, иначе помещать в отдельный бин
                'max_or_separate' - если доля пустых значений меньше missing_min_part, то объединять с бином с максимальным WOE, иначе помещать в отдельный бин
                'nearest_or_separate' - если доля пустых значений меньше missing_min_part, то объединять с ближайшим по WOE бином, иначе помещать в отдельный бин
        :param missing_min_part: минимальная доля пустых значений для выделения отдельного бина при missing_process 'min_or_separate', 'max_or_separate' или 'nearest_or_separate'

        ---Обработка значений, отсутствующих в биннинге---
        :param others_process: Способ обработки:
                'missing': отсутствующим значениям присваивается WOE, соответствующий пустым значениям
                'min': отсутствующим значениям присваивается минимальный WOE
                'max': отсутствующим значениям присваивается максимальный WOE
                float: отсутствующим значениям присваивается заданный фиксированный WOE
        :param opposite_sign_to_others: В случае, когда непрерывная переменная на выборке для разработки имеет только один знак,
                то все значения с противоположным знаком относить к others
        """
        self.feature = feature
        self.round_digits = round_digits
        self.round_woe = round_woe
        self.rounding_migration_coef = rounding_migration_coef
        self.simple = simple
        self.n_folds = n_folds
        self.woe_adjust = woe_adjust
        self.alpha = alpha
        self.alpha_recalc = self.alpha is None
        self.alpha_range = alpha_range
        self.alpha_scoring = alpha_scoring
        self.alpha_best_criterion = alpha_best_criterion
        self.missing_process = missing_process
        self.missing_min_part = missing_min_part
        self.others_process = others_process
        self.opposite_sign_to_others = opposite_sign_to_others
        self.special_bins = special_bins
        self.woe_threshold_missing = 0
        self.others_woe = np.nan
        self.is_active = False
        if feature in ds.cat_columns:
            self.categorical_type = str(ds.samples[ds.train_name][feature].dtype)
            self.groups = {0: ds.samples[ds.train_name][feature].unique().tolist()}
        else:
            self.categorical_type = ''
            self.groups = {0: [-np.inf, np.inf]} # for categorical - {group_number: [list of values]}, for ordered -{group_number: [left_bound, right_bound]}
        # -1 - group number for missings
        self.woes = {0: 0} # group_number:woe
        self.missing_group = -1
        self.history = []
        if ds.samples is not None:
            samples = {name: ds.samples[name][[ds.target, feature]] for name in ds.samples}
        else:
            samples = None        
        self.ds = DataSamples(samples=samples, target=ds.target, features=[feature], cat_columns=[feature] if feature in ds.cat_columns else [],
                              n_jobs=1, random_state=ds.random_state)
        if ds.bootstrap_base is not None:
            self.ds.bootstrap_base = ds.bootstrap_base[[ds.target, feature]]
            self.ds.bootstrap = ds.bootstrap

    def calc_simple_woes(self, df=None):
        '''
        Simply calculates regularized WOE for each interval.
        Formula for regularized WOE of the i-th interval (value group) of feature:
        SnoothedWOE_i = log((n + alpha)*DefaultRate/(n*DefaultRate_i + alpha)),
        where n is number of samples, DefaultRate = N_bad/N_good, DefaultRate_i = (N_bad_i+woe_adjust)/(N_good_i+woe_adjust).

        Returns
        ----------
        woes: {group: woe}
        '''
        groups_stat = self.calc_groups_stat(df)
        alpha = self.alpha if self.alpha is not None else 0
        groups_stat['woe'] = (np.log(groups_stat['n1'].sum() / groups_stat['n0'].sum() * (alpha + groups_stat['n']) /
                                       (groups_stat['n'] * (groups_stat['n1'] + self.woe_adjust) / (
                                               groups_stat['n0'] + self.woe_adjust) + alpha))).round(self.round_woe)
        return groups_stat['woe'].to_dict()

    def woe_folds(self):
        '''
        Breaks the feature into folds for each value group (interval) and calculates regularized WOE for each fold.
        Formula for regularized WOE of the i-th interval (value group) of feature:
        SmoothedWOE_i = log((n + alpha)*DefaultRate/(n*DefaultRate_i + alpha)),
        where n is number of samples, DefaultRate = N_bad/N_good, DefaultRate_i = N_bad_i/N_good_i.

        WOE for folds is simular to cross-validation:
        1. Samples of an interval (value group) are divided into n_folds folds
        2. For each (n_folds - 1) intervals (value groups) SmoothedWOE is calculated
        3. For each fold its SmoothedWOE is the SmoothedWOE calculated on the other folds

        Example:
        Feature 'AGE', interval '25-35 years old', n_folds = 5.
        All the clients with AGE between 25 and 35 years are divided into 5 folds. For clients of the 1st fold SmoothedWOE value is SmoothedWOE calculated on 2-5 folds.

        Returns
        ----------
        woes: {left_bound : {fold_num : woe}} where left_bound is a lower bound of an interval, fold_num is a number of a fold (from 0 to n_folds-1), woe is a WOE value for the fold
        folds: {left_bound : {fold_num : fold_indexes}} where left_bound is a lower bound of an interval, fold_num is a number of a fold (from 0 to n_folds-1),  fold_indexes is indexes of samples in the fold
        '''
        groups_stat = self.calc_groups_stat()
        DR = groups_stat['n1'].sum() / groups_stat['n0'].sum()

        folds = {}
        # calculation of folds
        for group, data_i in self.ds.samples[self.ds.train_name].groupby('group'):
            folds[group] = {}
            if data_i.shape[0] > self.n_folds:
                skf = StratifiedKFold(self.n_folds)
                tmp = 0
                for train_index, test_index in skf.split(X = data_i, y = data_i[self.ds.target]):
                    # indexes addition
                    folds[group][tmp] = [data_i.iloc[train_index].index, data_i.iloc[test_index].index]
                    tmp = tmp + 1
            else:
                folds[group][0] = [data_i.index, data_i.index]
        # WOE for each fold and interval
        woes = {}

        # for each interval
        for group, data_i in self.ds.samples[self.ds.train_name].groupby('group'):
            woes[group] = {}
            #for each fold
            for fold in folds[group]:
                indexes_fold = folds[group][fold][0]
                data_fold = data_i.loc[indexes_fold]
                N_b_i = data_fold[self.ds.target].sum()
                N_g_i = data_fold.shape[0] - N_b_i
                n = N_g_i + N_b_i
                if n != 0:
                    DR_i = (N_b_i + self.woe_adjust)/(N_g_i + self.woe_adjust)
                    n = N_g_i + N_b_i
                    smoothed_woe_i = round(np.log(DR*(self.alpha + n)/(n*DR_i + self.alpha)), self.round_woe)
                    woes[group][fold] = smoothed_woe_i

        #removing bounds with no data cooresponding to them (in case of empty dictionary for folds)
        woes={x:woes[x] for x in woes if woes[x]!={}}
        return woes, folds
        # woes: {group : {fold_num : woe}}
        # folds: {group : {fold_num : fold_indexes}}

    def calc_woe_folds(self):
        '''
        Calculates WOE for each sample according to folds

        Returns
        ----------
        woes - a dictionary {group: {fold_number: woe}},
        result - a list of values transformed to woe by folds
        '''
        if self.alpha is None:
            print ('Achtung bitte! Keine Alpha... Bis dann! :) ')
            return None
        woes, folds = self.woe_folds()
        # for each sample finds its interval (values group), fold and, consequently, WOE
        result = [woes[group][fold] for index, row in self.ds.samples[self.ds.train_name][[self.feature, self.ds.target]].iterrows()
                                    for group in folds
                                    for fold in folds[group]
                                    if index in folds[group][fold][1]]
        return woes, result

    def optimize_alpha(self):
        '''
        Optimal alpha selection for WoE-transformed data

        Returns
        --------
        optimal alpha

        '''
        if self.alpha_range is None:
            self.alpha_range = range(10, 100, 10)
        classifier = LogisticRegression(random_state=self.ds.random_state)
        scores = {}
        for alpha in self.alpha_range:
            self.alpha = alpha
            if self.simple:
                x = self.set_avg_woes(woes=self.calc_woes())
            else:
                x = self.calc_woe_folds()[1]
            scores[alpha] = np.mean(cross_val_score(classifier, x, self.ds.samples[self.ds.train_name][self.ds.target],
                                                    cv=5, scoring=self.alpha_scoring))

        if self.alpha_best_criterion == 'min':
            self.alpha = min(scores, key=scores.get)
        elif self.alpha_best_criterion == 'max':
            self.alpha = max(scores, key=scores.get)
        else:
            self.alpha = 0
        return self.alpha

    def sorted_groups(self, groups=None):
        if groups is None:
            groups = self.groups.copy()
        if self.categorical_type:
            return groups
        else:
            return {**{k: v for k, v in groups.items() if isinstance(v, list) and len(v) == 1}, **groups}

    def get_group_condition_code(self, group):
        vals = self.groups[group]
        data_s = f"df['{self.feature}']"
        if group == -1:
            return f"{data_s}.isnull()"
        if self.categorical_type:
            if self.missing_group != group:
                return f"{data_s}.isin({vals})"
            else:
                return f"({data_s}.isin({vals})) | ({data_s}.isnull())"
        if isinstance(vals, list):
            if len(vals) == 2:
                first = f'({data_s} >= {vals[0]})' if vals[0] != -np.inf else ''
                second = f'({data_s} < {vals[1]})' if vals[1] != np.inf else ''
                if not first and not second:
                    first = f"~{data_s}.isnull()"
                if self.missing_group != group:
                    return f"{first}{' & ' if first and second else ''}{second}"
                else:
                    return f"({first}{' & ' if first and second else ''}{second}) | ({data_s}.isnull())"
            else:
                return f"{data_s} == {vals[0]}"

    def get_transform_func(self, start=''):
        groups_list = [g for g in self.sorted_groups() if g in self.woes]
        s = start
        for i, group in enumerate(groups_list):
            s += f"np.where({self.get_group_condition_code(group)}, {self.woes[group]}, \n" + " " * (len(start) + 9 * (i + 1))
        s += f'{self.others_woe}{")" * len(groups_list)}'.replace('nan', 'np.nan')
        return s

    def get_condlist(self, data, groups):
        condlist = []
        groups = self.sorted_groups(groups)
        for group in groups:
            vals = groups[group]
            if group == -1:
                condlist.append(data.isnull())
            else:
                if self.categorical_type:
                    if self.missing_group != group:
                        condlist.append(data.isin(vals))
                    else:
                        condlist.append(data.isin(vals) | data.isnull())
                elif isinstance(vals, list):
                    if len(vals) == 2:
                        if self.missing_group != group:
                            condlist.append((data >= vals[0]) & (data < vals[1]))
                        else:
                            condlist.append(((data >= vals[0]) & (data < vals[1])) | (data.isnull()))
                    else:
                        condlist.append(data == vals[0])
                else:
                    condlist.append(data == vals)
        return condlist, groups

    def set_avg_woes(self, data=None, groups=None, woes=None):
        '''
        Replaces all values of a feature to related WOE

        Parameters
        -----------
        data: a Series, containing initial values of feature
        groups: a dictionary with groups description
        woes: a dictionary with WOEs for groups

        Returns
        -----------
        a Series of WOE-transformed feature values
        '''

        if data is None:
            data = self.ds.samples[self.ds.train_name][self.feature]
        if groups is None:
            groups = self.groups
        if woes is None:
            woes = self.woes
        condlist, groups = self.get_condlist(data, groups)
        return np.select(condlist, [woes[group] for group in groups], self.others_woe)

    def calc_groups_stat(self, df=None):
        groups_stat = (df if df is not None else self.ds.samples[self.ds.train_name]).groupby('group')\
                      .agg(n=(self.ds.target, 'count'), n1=(self.ds.target, 'sum'))
        groups_stat['n0'] = groups_stat['n'] - groups_stat['n1']
        groups_stat['woe'] = groups_stat.index.map(self.woes)
        return groups_stat

    def set_groups(self, data=None, groups=None, inplace=False):
        '''
        Replaces all values of a feature to related group

        Parameters
        -----------
        data: a Series, containing initial values of feature
        groups: a dictionary with groups description
        woes: a dictionary with WOEs for groups

        Returns
        -----------
        a Series of corresponding groups for input values
        '''
        if data is None:
            data = self.ds.samples[self.ds.train_name][self.feature]
        if groups is None:
            groups = self.groups
        condlist, groups = self.get_condlist(data, groups)
        result = np.select(condlist, list(groups.keys()), 404)
        if inplace:
            self.ds.samples[self.ds.train_name]['group'] = result
        return result

    def print_woe(self):
        '''
        Prints WOE parameters in a standard and convenient way
        '''
        groups_stat = self.calc_groups_stat().reset_index()
        groups_stat[['n', 'n1']] = groups_stat[['n', 'n1']].round(0).astype('int')
        groups_stat['values'] = groups_stat['group'].map(self.groups).astype('str')
        if self.missing_group != -1 and not self.categorical_type:
            groups_stat.loc[groups_stat['group'] == self.missing_group, 'values'] += ' + missings'
        for i, group in enumerate(self.special_bins.keys()):
            groups_stat.loc[groups_stat['group'] == -2 - i, 'group'] = group
        print('\nCurrent binning:')
        print(groups_stat[['group', 'values', 'woe', 'n', 'n1']].to_string(index=False))
        print()

    def BusinessLogicChecker(self, BL_allow_Vlogic_to_increase_gini=100, verbose=False):
        '''
        Checks if the business logic condition is True

        Parameters
        -----------
        BL_allow_Vlogic_to_increase_gini:
        verbose: if comments and graphs should be printed

        Returns
        ----------
        Boolean - whether the check was successful and dataframe of check log
        '''
        if verbose > 0:
            print('\n------------- Business logic checks --------------')
        if not self.categorical_type:
            woes_dropna = {self.groups[x][0]: self.woes[x] for x in self.woes if
                           isinstance(self.groups[x], list) and len(self.groups[x]) == 2}
            groups_info = pd.DataFrame(woes_dropna, index=['woe']).transpose().reset_index().rename({'index': 'lower'},
                                                                                                    axis=1)
            groups_info['upper'] = groups_info['lower'].shift(-1).fillna(np.inf)
            if groups_info.shape[0] == 1:
                if verbose > 0:
                    print('Only one group with non-missing values is present. Skipping trend check...')
                trend = ''
                trend_type = 'one group'
            else:
                gi_check = groups_info.dropna(how='all', subset=['lower', 'upper'])[['woe', 'lower', 'upper']].copy()
                gi_check['risk_trend'] = np.sign((gi_check['woe'] - gi_check['woe'].shift(1)).dropna()).apply(lambda x: '+' if (x > 0) else '-' if (x < 0) else '0')
                trend = gi_check['risk_trend'].str.cat()
                if re.fullmatch('-*|\+*', trend):
                    trend_type = 'monotonic'
                    if verbose > 0:
                        print(f'WOE trend is monotonic')
                elif BL_allow_Vlogic_to_increase_gini < 100 and re.fullmatch('-*\+*|\+*-*', trend):
                    gini = self.ds.calc_gini(features=[add_suffix(self.feature)], samples=[self.ds.train_name]).at[add_suffix(self.feature), self.ds.train_name]
                    if gini < BL_allow_Vlogic_to_increase_gini:
                        trend_type = 'no trend'
                        if verbose > 0:
                            print(f'WOE trend is V-shaped, but gini is too low ({gini} < {BL_allow_Vlogic_to_increase_gini})')
                    else:
                        trend_type = 'V-shape'
                        if verbose > 0:
                            print(f'WOE trend is V-shaped')
                else:
                    trend_type = 'no trend'
                    if verbose > 0:
                        print(f'WOE trend does not have right shape')
            if verbose > 0:
                fig = plt.figure(figsize=(5, 0.5))
                plt.plot(range(len(groups_info.dropna(how='all', subset=['lower', 'upper'])['lower'])),
                         groups_info.dropna(how='all', subset=['lower', 'upper'])['woe'], color='red')
                plt.xticks(range(len(groups_info.dropna(how='all', subset=['lower', 'upper'])['lower'])),
                           round(groups_info.dropna(how='all', subset=['lower', 'upper'])['lower'], 3), fontsize=8)
                plt.ylabel('WOE', fontsize=8)
                plt.yticks(fontsize=8)
                fig.autofmt_xdate()
                plt.show()
        else:
            trend = ''
            trend_type = 'categorical'
            if verbose > 0:
                print('Categorical feature. Skipping trend check...')
        if verbose > 0:
            print('...Passed!' if trend_type != 'no trend' else '...Failed!')
        return trend_type, [self.feature, 0, trend, trend_type]
    
    def GiniChecker(self, gini_threshold=5, gini_decrease_threshold=0.2, gini_increase_restrict=True,
             verbose=False, with_test=False):
        '''
        Checks if gini of the feature is significant and stable enough

        Parameters
        -----------
        gini_threshold: gini on train and validate/95% bootstrap should be greater then this
        gini_decrease_threshold: gini decrease from train to validate/95% bootstrap deviation from mean to mean should be greater then this
        gini_increase_restrict: if gini increase should also be restricted
        verbose: if comments and graphs should be printed
        with_test: add checking of gini values on test (calculation is always on)

        Returns
        ----------
        Boolean - whether the check was successful and dictionary of gini values for all available samples
        '''
        if verbose > 0:
            print('\n------------------ Gini checks -------------------')

        gini_correct = True
        f_WOE = add_suffix(self.feature)
        for name in self.ds.samples:
            if name != self.ds.train_name:
                self.ds.samples[name][f_WOE] = self.set_avg_woes(data=self.ds.samples[name][self.feature])
        gini_values = self.ds.calc_gini(features=[f_WOE], samples=list(self.ds.samples)).T[f_WOE].to_dict()
        for name, gini in gini_values.items():
            if gini < gini_threshold and (name == self.ds.train_name or with_test):
                gini_correct = False
                if verbose > 0:
                    print(f'Gini {name} is less then threshold {gini_threshold}')
            if name != self.ds.train_name and with_test:
                if gini_decrease_threshold < 1:
                    decrease = 1 - gini / gini_values[self.ds.train_name]
                else:
                    decrease = gini_values[self.ds.train_name] - gini
                if decrease > gini_decrease_threshold or (gini_increase_restrict and -decrease > gini_decrease_threshold):
                    gini_correct = False
                    if verbose > 0:
                        print(f'Gini change from {self.ds.train_name} to {name} is greater then threshold: {round(decrease, self.round_woe)} > {gini_decrease_threshold}')
        if self.ds.bootstrap_base is not None and gini_correct:
            self.ds.bootstrap_base[f_WOE] = self.set_avg_woes(data=self.ds.bootstrap_base[self.feature])
            gini_df = self.ds.calc_gini(features=[f_WOE], samples=['Bootstrap'])
            gini_values['Bootstrap mean'] = gini_df.at[f_WOE, 'Bootstrap mean']
            gini_values['Bootstrap std'] = gini_df.at[f_WOE, 'Bootstrap std']
            gini_5 = gini_values['Bootstrap mean'] - 1.96*gini_values['Bootstrap std']
            if gini_5 < gini_threshold:
                gini_correct = False
                if verbose > 0:
                    print(f'Less then 95% of gini distribution is greater then threshold: (mean-1.96*std) {round(gini_5, self.round_woe)} < {gini_threshold}')
            if gini_decrease_threshold < 1:
                decrease = 1.96*gini_values['Bootstrap std']/gini_values['Bootstrap mean']
            else:
                decrease = 1.96*gini_values['Bootstrap std']
            if decrease > gini_decrease_threshold:
                gini_correct = False
                if verbose > 0:
                    if gini_decrease_threshold < 1:
                        print(f'Gini deviation from mean for 95% of distribution is greater then threshold: (1.96*std/mean) {round(decrease,self.round_woe)} > {gini_decrease_threshold}')
                    else:
                        print(f'Gini deviation for 95% of distribution is greater then threshold: (1.96*std) {round(decrease, self.round_woe)} > {gini_decrease_threshold}')
        if verbose > 0:
            print(pd.DataFrame(gini_values, index=['Gini']).round(self.round_woe).to_string())
            print('...Passed!' if gini_correct else '...Failed!')
        return gini_correct, gini_values
    
    def WOEOrderChecker(self, dr_threshold=0.01, correct_threshold=0.85, woe_adjust=0.5, miss_is_incorrect=True,
                        with_test=False, verbose=False):
        '''
        Checks if WoE order of the feature remains stable in bootstrap

        Parameters
        -----------
        dr_threshold: if WoE order is not correct, then default rate difference between swaped bins is checked
        correct_threshold: what part of checks on bootstrap should be correct for feature to pass the check
        woe_adjust: woe adjustment factor (for Default_Rate_i formula)
        miss_is_incorrect: is there is no data for a bin on bootstrap sample, should it be treated as error or not
        with_test: check order on all samples
        verbose: if comments and graphs should be printed

        Returns
        ----------
        Boolean - whether the check was successful and dataframes with WoE and ER values for groups per existing sample
        '''

        def calc_cur_WOE_ER(df):
            groups_stat = self.calc_groups_stat(df=df)
            groups_stat['woe'] = (np.log(groups_stat['n1'].sum() / groups_stat['n0'].sum() * (self.alpha + groups_stat['n'])
                                          / (groups_stat['n'] * (groups_stat['n1'] + self.woe_adjust) / (groups_stat['n0'] + self.woe_adjust) + self.alpha))).round(self.round_woe)
            groups_stat['ER'] = groups_stat['n1'] / groups_stat['n']
            return groups_stat['woe'].to_dict(), groups_stat['ER'].to_dict()

        def correct_order(cur_sample_woe, group_er, miss_is_incorrect):
            cur_sample_woe['er'] = cur_sample_woe['group'].map(group_er)
            cur_sample_woe['trend_train'] = np.sign((cur_sample_woe[self.ds.train_name] - cur_sample_woe[self.ds.train_name].shift(1)))
            cur_sample_woe['trend'] = np.sign((cur_sample_woe[name] - cur_sample_woe[name].shift(1)))
            cur_sample_woe['big_delta_er'] = (cur_sample_woe['er'] - cur_sample_woe['er'].shift(1)).abs() > dr_threshold
            if cur_sample_woe[(cur_sample_woe['trend_train'] != cur_sample_woe['trend']) & (cur_sample_woe['big_delta_er'])].dropna(subset=['trend_train', 'trend']).shape[0] > 0 or (miss_is_incorrect and cur_sample_woe[name].isnull().any()):
                return 0
            return 1

        if verbose > 0:
            print('\n---------------- WOE order checks ----------------')
        tmp = self.ds.samples[self.ds.train_name].copy()
        tmp = tmp[tmp['group'] >= -1]
        groups = sorted(tmp['group'].unique())
        out_er = tmp[['group', self.ds.target]].groupby('group', as_index=False).mean().rename({self.ds.target: self.ds.train_name}, axis=1)
        out_woes = out_er.copy()
        out_woes[self.ds.train_name] = out_woes['group'].map(self.woes)
        for name, sample in self.ds.samples.items():
            if name == self.ds.train_name:
                continue
            cur_sample = sample[[self.feature, self.ds.target]]
            cur_sample['group'] = self.set_groups(data=cur_sample[self.feature])
            group_woe, group_er = calc_cur_WOE_ER(cur_sample)
            out_woes[name] = out_woes['group'].map(group_woe)
            out_er[name] = out_er['group'].map(group_er)
            if with_test:
                if not correct_order(out_woes[['group', self.ds.train_name, name]].copy(), group_er, miss_is_incorrect):
                    if verbose > 0:
                        print(f'...Failed!/nTrend breaking on sample {name}')
                    return False, out_woes, out_er
                elif verbose > 0:
                    print(f'Trend holding on sample {name}')
        if self.ds.bootstrap_base is not None:
            bootstrap_correct = []
            self.ds.bootstrap_base['group'] = self.set_groups(data=self.ds.bootstrap_base[self.feature])
            for bn, idx in enumerate(self.ds.bootstrap):
                cur_sample = self.ds.bootstrap_base.iloc[idx]
                name = f'Bootstrap {bn}'
                group_woe, group_er = calc_cur_WOE_ER(cur_sample)
                out_woes[name] = out_woes['group'].map(group_woe)
                out_er[name] = out_er['group'].map(group_er)
                bootstrap_correct.append(correct_order(out_woes[['group', self.ds.train_name, name]].copy(), group_er, miss_is_incorrect))
            result = sum(bootstrap_correct)/len(bootstrap_correct) >= correct_threshold
            if verbose > 0:
                print(f'Trend holding on {sum(bootstrap_correct)}/{len(bootstrap_correct)} bootstrap samples, threshold = {correct_threshold}')
                print('...Passed!' if result else '...Failed!')
            return result, out_woes, out_er
        else:
            if verbose > 0:
                print('No bootstrap samples were found is self.ds object/n...Passed!.')
            return True, out_woes, out_er

    def auto_fit(self, verbose=True, method='tree', max_n_bins=10, min_bin_size=0.05,
                 criterion='entropy', scoring='neg_log_loss', max_depth=None, solver='cp', divergence='iv',
                 WOEM_on=True, WOEM_woe_threshold=0.05, WOEM_with_missing=False,
                 SM_on=True, SM_target_threshold=5, SM_size_threshold=100,
                 BL_on=True, BL_allow_Vlogic_to_increase_gini=100,
                 G_on=True, G_gini_threshold=5, G_with_test=False, G_gini_decrease_threshold=0.2, G_gini_increase_restrict=True,
                 WOEO_on=False, WOEO_dr_threshold=0.01, WOEO_correct_threshold=0.85, WOEO_miss_is_incorrect=True, WOEO_with_test=False
                 ):
        '''
        Attempts to find suitable binning satisfying all conditions and passing all checks, adjusted by user parameters
        Values for groups amount are taken from groups_range in the specified order, feature is being fitted,
        then gini, business logic and WoE order checks are ran. If all checks are passed, then current binning stays,
        otherwise the feature is fitted with the next value of groups amount and checks are carried out again.

        Parameters
        -----------
        max_n_bins: list of integers that will be used as max_leaf_nodes value in FeatureWOE.fit method in the specified order
        verbose: if comments and graphs from checks should be printed

        Fit options:
        -----------
        scoring: a measure for cross-validation used used for optimal WOE splitting
        max_depth: the maximum of the DecisionTree depth used for optimal WOE splitting
        min_samples_leaf: the minimum of the DecisionTree leaf size used for optimal WOE splitting (one value)

        Size merge options (SM):
        -----------------------
        SM_on: flag to turn on/off WoE merge process
        SM_target_threshold: min number of targets for group to not be considered small
        SM_size_threshold: min number of observations for group to not be considered small

        WoE merge options (WOEM):
        -----------------------
        WOEM_on: flag to turn on/off merging by WoE process
        WOEM_woe_threshold: if woe difference between groups (neighboring groups for interval) is less then this threshold, then they are to be merged
        WOEM_with_missing: should woe difference with missing group also be checked

        Business logic check options (BL):
        ---------------------------------
        BL_on: flag to turn on/off business logic check
        BL_allow_Vlogic_to_increase_gini:

        WoE order check options (WOEO):
        ------------------------------
        WOEO_on: flag to turn on/off WoE order check
        WOEO_dr_threshold: if WoE order is not correct, then default rate difference between swaped bins is checked
        WOEO_correct_threshold: what part of checks on bootstrap should be correct for feature to pass the check
        WOEO_woe_adjust: woe adjustment factor (for Default_Rate_i formula)
        WOEO_miss_is_incorrect: is there is no data for a bin on bootstrap sample, should it be treated as error or not

        Gini check options (G):
        ----------------------
        G_on: flag to turn on/off Gini check
        G_gini_threshold: gini on train and validate/95% bootstrap should be greater then this
        G_gini_decrease_threshold: gini decrease from train to validate/95% bootstrap deviation from mean to mean should be greater then this
        G_gini_increase_restrict: if gini increase should also be restricted
        G_with_test: should features be also checked on test sample (and be excluded in case of failure)

        Returns
        ----------
        A boolean value: True, if successful binning was found, else False
        and dataframes with log, gini, business logic, WoE and ER information for export
        '''
        def checks_to_dfs(checks):
            # log
            checks[0] = pd.DataFrame(checks[0], columns=['feature', 'iteration', 'n_bins', 'result', 'reason'])
            # Business Logic
            checks[1] = pd.DataFrame(checks[1], columns=['feature', 'iteration', 'trend', 'trend_type'])
            # Gini checks
            if checks[2]:
                checks[2] = pd.DataFrame(checks[2])
            else:
                checks[2] = pd.DataFrame()
            # WoE checks, ER checks, scorecards
            for i in [3, 4, 5]:
                if checks[i]:
                    checks[i] = pd.concat(checks[i])
                else:
                    checks[i] = pd.DataFrame()
            return checks

        f_WOE = add_suffix(self.feature)
        checks = [[] for i in range(6)]
        gini_V = 0
        check_dfs_V = [[] for i in range(6)]
        n_bins_fact = max_n_bins + 1
        iteration = 0
        if verbose > 0:
            print(f"\n{f' Auto binning for {self.feature} ':-^100}\n")
        self.woe_threshold_missing = WOEM_woe_threshold if WOEM_on and WOEM_with_missing else 0
        for n_bins in range(max_n_bins, 1, -1):
            if n_bins >= n_bins_fact:
                continue
            iteration +=1
            if verbose > 0:
                print(f"\n{f' Searching for the best split into {n_bins} groups ':-^75}\n")
            try:
                self.fit(new_groups=True, to_history=False, verbose=verbose, method=method, max_n_bins=n_bins,
                         min_bin_size=min_bin_size, monotonic=BL_on and (BL_allow_Vlogic_to_increase_gini == 100) or (gini_V > 0),
                         criterion=criterion, scoring=scoring, max_depth=max_depth, solver=solver, divergence=divergence)
            except Exception as e:
                if self.categorical_type:
                    values = self.ds.samples[self.ds.train_name][self.feature].dropna().unique().tolist()
                    try:
                        self.groups = {0: sorted(values)}
                    except:
                        self.groups = {0: values}
                else:
                    self.groups = {0: [-np.inf, np.inf]}
                self.fit(new_groups=False)
                if verbose != -1:
                    print(f'Exception! Feature {self.feature}, fit: {e}')
                checks[0].append([self.feature, iteration, n_bins, 'Exception', f'Fit: {e}'])
                checks[2].append({'feature': self.feature, 'iteration': iteration, self.ds.train_name: self.ds.calc_gini(features=[f_WOE], samples=[self.ds.train_name]).values[0][0]})
                continue

            n_bins_fact = len([x for x in self.woes if not pd.isnull(self.woes[x]) and x != -1 and not isinstance(x, str)])
            if SM_on:
                try:
                    self.merge_by_size(target_threshold=SM_target_threshold, size_threshold=SM_size_threshold, verbose=verbose)
                except Exception as e:
                    if verbose != -1:
                        print(f'Exception! Feature {self.feature}, Merging by size: {e}')

            if WOEM_on:
                try:
                    self.merge_by_woe(woe_threshold=WOEM_woe_threshold, verbose=verbose)
                except Exception as e:
                    if verbose != -1:
                        print(f'Exception! Feature {self.feature}, Merging by WOE: {e}')

            checks[5].append(self.export_scorecard(iteration=iteration))
            if len([x for x in self.woes if not pd.isnull(self.woes[x])]) == 1:
                if verbose > 0:
                    print(f'After the attempt with {n_bins} groups only one group remains. Continue cycle...')
                checks[0].append([self.feature, iteration, n_bins_fact, 'Failure', 'After merging only one group remains'])
                checks[2].append({'feature': self.feature, 'iteration': iteration, self.ds.train_name: self.ds.calc_gini(features=[f_WOE], samples=[self.ds.train_name]).values[0][0]})
                continue
            self.to_history(dupl=False)

            if BL_on:
                try:
                    BL_trend_type, BL_check = self.BusinessLogicChecker(BL_allow_Vlogic_to_increase_gini=BL_allow_Vlogic_to_increase_gini if gini_V == 0 else 100, verbose=verbose)
                except Exception as e:
                    if verbose != -1:
                        print(f'Exception! Feature {self.feature}, Business logic checks: {e}')
                    checks[0].append([self.feature, iteration, n_bins_fact, 'Exception', f'Business logic checks: {e}'])
                    checks[2].append({'feature': self.feature, 'iteration': iteration, self.ds.train_name: self.ds.calc_gini(features=[f_WOE], samples=[self.ds.train_name]).values[0][0]})
                    continue
                BL_check[1] = iteration
                checks[1].append(BL_check)
                if BL_trend_type == 'no trend':
                    checks[0].append([self.feature, iteration, n_bins_fact, 'Failure', 'Business logic check failed'])
                    checks[2].append({'feature': self.feature, 'iteration': iteration, self.ds.train_name: self.ds.calc_gini(features=[f_WOE], samples=[self.ds.train_name]).values[0][0]})
                    continue

            if G_on:
                try:
                    correct, gini_values = self.GiniChecker(gini_threshold=G_gini_threshold,
                           gini_decrease_threshold=G_gini_decrease_threshold,
                           gini_increase_restrict=G_gini_increase_restrict, verbose=verbose, with_test=G_with_test)
                except Exception as e:
                    if verbose != -1:
                        print(f'Exception! Feature {self.feature}, Gini checks: {e}')
                    checks[0].append([self.feature, iteration, n_bins_fact, 'Exception', f'Gini checks: {e}'])
                    continue
                checks[2].append({**{'feature': self.feature, 'iteration': iteration}, **gini_values})
                if not correct:
                    checks[0].append([self.feature, iteration, n_bins_fact, 'Failure', 'Gini check failed'])
                    continue
            else:
                gini_values = self.ds.calc_gini(features=[f_WOE], samples=[self.ds.train_name]).T[f_WOE].to_dict()
                checks[2].append({'feature': self.feature, 'iteration': iteration, self.ds.train_name: gini_values[self.ds.train_name]})

            if WOEO_on:
                try:
                    correct, add_woe, add_er = self.WOEOrderChecker(dr_threshold=WOEO_dr_threshold,
                            correct_threshold=WOEO_correct_threshold, woe_adjust=self.woe_adjust,
                            miss_is_incorrect=WOEO_miss_is_incorrect, with_test=WOEO_with_test, verbose=verbose)
                except Exception as e:
                    if verbose != -1:
                        print(f'Exception! Feature {self.feature}, WOE order checks: {e}')
                    checks[0].append([self.feature, iteration, n_bins_fact, 'Exception', f'WOE order checks: {e}'])
                    continue
                add_woe.insert(loc=0, column='feature', value=self.feature)
                add_woe.insert(loc=1, column='iteration', value=iteration)
                add_er.insert(loc=0, column='feature', value=self.feature)
                add_er.insert(loc=1, column='iteration', value=iteration)
                checks[3].append(add_woe)
                checks[4].append(add_er)
                if not correct:
                    checks[0].append([self.feature, iteration, n_bins_fact, 'Failure', 'WOE order check failed'])
                    continue

            if BL_on and BL_trend_type == 'V-shape':
                checks[0].append([self.feature, iteration, n_bins_fact, 'Success', 'Business logic is V-shaped, keep iterating to compare with the monotonic trend'])
                if gini_values[self.ds.train_name] > gini_V:
                    gini_V = gini_values[self.ds.train_name]
                    fact_number_groups_V = n_bins_fact
                    for i in range(6):
                        if len(checks[i]) > 0:
                            check_dfs_V[i] = checks[i][-1].copy()
                if verbose > 0:
                    print('Business logic is V-shaped, keep iterating to compare with the monotonic trend.')
                continue

            checks[0].append([self.feature, iteration, n_bins_fact, 'Success', ''])
            if gini_V > 0:
                gini_delta = gini_V - gini_values[self.ds.train_name]
                if gini_delta > BL_allow_Vlogic_to_increase_gini:
                    iteration = iteration + 1
                    check_dfs_V[0] = [self.feature, iteration, fact_number_groups_V, 'Success', f'V-shaped binning is chosen, because it gives an increase gini by {round(gini_delta, self.round_woe)}']
                    for i in range(6):
                        if len(check_dfs_V[i]) > 0:
                            if i in [0, 1]:
                                check_dfs_V[i][1] = iteration
                            else:
                                check_dfs_V[i]['iteration'] = iteration
                            checks[i].append(check_dfs_V[i])
                    if verbose > 0:
                        print(f'Finally V-shaped binning with {fact_number_groups_V} groups is chosen, because it gives an increase gini by {round(gini_delta, self.round_woe)}')
            return True, checks_to_dfs(checks)
        else:
            if gini_V > 0:
                iteration = iteration + 1
                check_dfs_V[0] = [self.feature, iteration, fact_number_groups_V, 'Success', 'V-shaped binning is chosen, because other binnings do not pass checks']
                for i in range(6):
                    if len(check_dfs_V[i]) > 0:
                        if i in [0, 1]:
                            check_dfs_V[i][1] = iteration
                        else:
                            check_dfs_V[i]['iteration'] = iteration
                        checks[i].append(check_dfs_V[i])
                if verbose > 0:
                    print(f'Finally V-shaped binning with {fact_number_groups_V} groups is chosen, because other binnings do not pass checks')
                return True, checks_to_dfs(checks)
            if verbose > 0:
                print('After all attempts no suitable binning was found.')
            return False, checks_to_dfs(checks)

    def set_round_groups(self, with_acceptable_migration=False, verbose=False):
        '''
        Rounds boundaries of groups. Checks if the rounding parameter is valid and extends it if necessary.
        Checks if groups do not collide and if groups' samples remain stable.

        Parameters
        -----------
        with_acceptable_migration: is it ok to allow migration between groups after rounding
        '''

        if not self.categorical_type:
            if with_acceptable_migration:
                # min interval between groups' boundariesf
                min_diff_b = min([1 if k < 0
                                  else (v[1] - self.ds.samples[self.ds.train_name][self.feature].min() if -np.inf in v
                                        else (self.ds.samples[self.ds.train_name][self.feature].max() - v[0] if np.inf in v
                                              else v[1] - v[0])) for (k, v) in self.groups.items()])
                change_rounds = False
                while min_diff_b < (.1)**self.round_digits:
                    self.round_digits += 1
                    change_rounds = True

                # checks changes in groups' volumes in case of rounding
                change_rounds_2 = True
                while change_rounds_2 and len(self.groups) > 1:
                    change_rounds_2 = False
                    for group in self.groups:
                        if group != -1:
                            left_from=round(self.groups[group][0], self.round_digits) if round(self.groups[group][0], self.round_digits) <= self.groups[group][0] else self.groups[group][0]
                            left_to=self.groups[group][0] if round(self.groups[group][0], self.round_digits) <= self.groups[group][0] else round(self.groups[group][0], self.round_digits)

                            right_from=round(self.groups[group][1], self.round_digits) if round(self.groups[group][1], self.round_digits) <= self.groups[group][1] else self.groups[group][1]
                            right_to=self.groups[group][1] if round(self.groups[group][1], self.round_digits) <= self.groups[group][1] else round(self.groups[group][1], self.round_digits)

                            migration = sum(self.ds.samples[self.ds.train_name][self.feature].apply(
                                                lambda x: (left_from <= x < left_to) or (right_from <= x < right_to)))
                            if migration/self.ds.samples[self.ds.train_name].shape[0] >= self.rounding_migration_coef and not change_rounds_2:
                                change_rounds_2 = True
                                self.round_digits += 1

                # rounding
                rounded_groups = {}
                for (k, v) in self.groups.items():
                    if k >= 0:
                        rounded_groups[k] = [round(v[0], self.round_digits), round(v[1], self.round_digits)]
                    else:
                        rounded_groups[k] = v
                if verbose > 0:
                    if change_rounds:
                        print ('The rounding parameter is too large, setting to', self.round_digits)
                self.groups = rounded_groups
            else:
                exact_edges=[]
                rounded_edges=[]
                for (k, v) in self.groups.items():
                    if k>0:
                        exact_edges.append(v[0])
                        before_split = self.ds.samples[self.ds.train_name][self.feature][self.ds.samples[self.ds.train_name][self.feature]<v[0]].max()
                        rounded_split = (before_split+v[0])/2
                        precision = len(str(rounded_split).split('.')[1])
                        previous_rounded_split = None
                        while rounded_split>before_split and rounded_split<v[0] and previous_rounded_split!=rounded_split:
                            previous_rounded_split=rounded_split
                            final_precision = precision
                            precision-=1
                            rounded_split = int(rounded_split) if precision==0 \
                                                else int((rounded_split)*(10**precision))/(10**precision)
                        candidate_split=int((before_split+v[0])/2) if final_precision==0 \
                                                 else int(((before_split+v[0])/2)*(10**final_precision))/(10**final_precision)

                        if final_precision<len(str(v[0]).split('.')[1])-(len(str(v[0]).replace('.',''))-len(str(v[0]).replace('.','').rstrip('0'))) and \
                           (self.ds.samples[self.ds.train_name][self.feature]<v[0]).sum()==(self.ds.samples[self.ds.train_name][self.feature]<candidate_split).sum():
                            rounded_edges.append(candidate_split)
                        else:
                            rounded_edges.append(v[0])

                rounded_groups = {}
                for (k, v) in self.groups.items():
                    if k >= 0:
                        rounded_groups[k] = [-np.inf if v[0]==-np.inf else rounded_edges[exact_edges.index(v[0])],
                                              np.inf if v[1]==np.inf  else rounded_edges[exact_edges.index(v[1])]]
                    else:
                        rounded_groups[k] = v

                if rounded_edges!=exact_edges:
                    if verbose > 0:
                        print ('Rounding edges:', str([exact_edges[i] for i in range(len(exact_edges)) if exact_edges[i]!=rounded_edges[i]]),
                               'to', str([rounded_edges[i] for i in range(len(rounded_edges)) if exact_edges[i]!=rounded_edges[i]]))
                    self.groups = rounded_groups

    def calc_woes(self, simple=False):
        if self.simple or simple:
            return self.calc_simple_woes()
        else:
            try:
                woe_folds = self.calc_woe_folds()[0]
                return {group: round(np.mean(np.array([woe_folds[group][fold] for fold in woe_folds[group]])), self.round_woe) for group in woe_folds}
            except ValueError:
                print ('ValueError for WOE calculation. Please check n_folds parameter and group sizes. Turning to simple WOE calculation... Hope it works.')
                return self.calc_simple_woes()

    def woe_fit(self, verbose=False, missings_to_process=True):
        '''
        Calculates WOE for FeatureWOE due to simple parameter
        '''
        # optimal alpha calculation
        if self.alpha_recalc:
            self.alpha = self.optimize_alpha()
            if verbose > 0:
                print(f'Optimal alpha: {self.alpha}')
        woes = self.calc_woes()
        if missings_to_process:
            woes_no_miss = copy.deepcopy(woes)
            woe_miss = woes_no_miss[-1]
            woes_no_miss = {k: v for k, v in woes_no_miss.items() if k >= 0}
            if self.missing_process == 'min':
                self.missing_group = min(woes_no_miss, key=woes_no_miss.get)
            elif self.missing_process == 'max':
                self.missing_group = max(woes_no_miss, key=woes_no_miss.get)
            elif self.missing_process == 'nearest' or any([abs(w - woe_miss) < self.woe_threshold_missing for w in woes_no_miss.values()]):
                nearest_group, nearest_woe = self.find_nearest(woe_miss, woes_no_miss)
                self.missing_group = nearest_group
            if self.missing_group != -1:
                del self.groups[-1]
                self.set_groups(inplace=True)
                woes = self.calc_woes()

        if self.others_process in ['missing_or_min', 'missing_or_max']:
            if self.missing_group in woes:
                self.others_woe = woes[self.missing_group]
            elif self.others_process == 'missing_or_min':
                self.others_woe = min(woes.values())
            elif self.others_process == 'missing_or_max':
                self.others_woe = max(woes.values())
        elif self.others_process == 'min':
            self.others_woe = min(woes.values())
        elif self.others_process == 'max':
            self.others_woe = max(woes.values())
        else:
            self.others_woe = self.others_process
        self.woes = woes
        return woes

    def categorical_to_interval(self, data):
        '''
        Transforms categorical features into interval oned via WOE calculation for each category.
        '''
        # WOE for each category: optimal alpha - woe by folds - final woe calculation
        # turning categorical values into separate groups for pre-woe
        self.groups = {group: [value] for group, value in enumerate(data.dropna().unique())}
        self.set_groups(inplace=True)
        self.woes = self.calc_woes(simple=True)
        return self.set_avg_woes(data=data)

    def get_tree_splits(self, dtree):
        '''
        Returns list of thresholds of the deision tree.

        Parameters
        ---------------
        tree: DecisionTreeClassifier object

        Returns
        ---------------
        boundaries: list of thresholds
        '''
        children_left = dtree.tree_.children_left
        children_right = dtree.tree_.children_right
        threshold = dtree.tree_.threshold
        # boundaries of split
        boundaries = [-np.inf, np.inf]

        # determination of groups
        for i in range(dtree.tree_.node_count):
            if children_left[i] != children_right[i]:
                boundaries.append(threshold[i])

        return sorted(boundaries)

    def find_nearest(self, x, values):
        '''
        Finds in values the nearest one for x. If 'values' is dict then returns the nearest value and its key and
        if 'values' is list then returns the nearest value and its index

        Parameters
        --------------
        x: float, the value to process
        values: dict or list of possible nearest values
        '''
        if isinstance(values, dict):
            diff = {abs(v-x): k for (k, v) in values.items()}
            return diff[min(diff)], values[diff[min(diff)]]
        elif isinstance(values, list) or isinstance(values, np.ndarray):
            diff = abs(np.array(values) - x)
            return diff.index(min(diff)), values[diff.index(min(diff))]

    def fit_gridsearch(self, x_train, y_train, parameters_grid, criterion, scoring):
        '''
        TECH

        Searches for the best decision tree for groups. Used in self.fit().

        Parameters
        ------------
        x_train: pd.Series to fit on
        y_train: pd.Series with targets for fit
        parameters_grid: parameters for gridsearch

        Returns
        ------------
        The best decision tree
        '''

        gridsearch = GridSearchCV(DecisionTreeClassifier(criterion=criterion, random_state=self.ds.random_state),
                                  parameters_grid, scoring=scoring, cv=7)
        gridsearch.fit(x_train[:, None], y_train)
        return gridsearch.best_estimator_

    def categorical_recover(self, pre_groups):
        '''
        Recovers self.groups and self.woes for categorical non-predefined features because such features are pre-processed
        '''
        if self.categorical_type:
            final_groups = {}
            for group, vals in self.groups.items():
                if isinstance(vals, list):
                    for category, pre_woe in self.woes.items():
                        if category in pre_groups and not pd.isnull(pre_woe) and pre_woe >= vals[0] and pre_woe < vals[1]:
                            if pd.isnull(pre_groups[category]):
                                self.missing_group = group
                            else:
                                if group in final_groups:
                                    final_groups[group] = final_groups[group] + pre_groups[category]
                                else:
                                    final_groups[group] = pre_groups[category]
            try:
                self.groups = {g: sorted(v) for g,v in final_groups.items()}
            except:
                self.groups = final_groups
            self.woes = {group: self.woes[group] for group in self.groups if group in self.woes}

    def fit(self, new_groups=True, to_history=True, verbose=False, method='tree', max_n_bins=10, min_bin_size=0.05,
            criterion='entropy', scoring='neg_log_loss', max_depth=None, monotonic=False, solver='cp', divergence='iv'):
        '''
        Optimizes alpha, determines optimal split into WOE intervals and calculates WOE. After that, the class is ready to
        transform unknown datasets containing the feature.

        Parameters
        -----------
        new_groups: if True, the bounds of the feature woe groups are deleted; usefull for refitting a Feature_woe object
        scoring: a measure for cross-validation used used for optimal WOE splitting
        max_leaf_nodes: the maximum of the DecisionTree leaves number used for optimal WOE splitting
        max_depth: the maximum of the DecisionTree depth used for optimal WOE splitting
        min_samples_leaf: the minimum of the DecisionTree leaf size used for optimal WOE splitting (one value)
        '''
        self.is_active = True
        self.missing_group = -1
        missings_to_process = self.ds.samples[self.ds.train_name][self.feature].isnull().any()
        if missings_to_process and self.missing_process.endswith('_or_separate'):
            if self.ds.samples[self.ds.train_name][self.feature].isnull().mean() < self.missing_min_part:
                self.missing_process = self.missing_process.rsplit('_or_separate', maxsplit=1)[0]
            else:
                self.missing_process = 'separate'
        if max_n_bins is None:
            max_n_bins = 10
        if min_bin_size is None:
            min_bin_size = 0.05
        if max_depth is None:
            max_depth = 5
        #optimal bounds calculation
        if new_groups:
            # For categorical features:
            # intermediate WOEs are calculated for each category => the features turns from categorical to numerical => for the further transformations the feature is considered numerical
            df = self.ds.samples[self.ds.train_name].copy()
            if missings_to_process:
                df = df[~df[self.feature].isnull()]
            if self.special_bins:
                df = df[~df[self.feature].isin(self.special_bins.values())]
            x_train = df[self.feature]
            y_tran = df[self.ds.target]
            if method == 'tree':
                if self.categorical_type:
                    x_train = self.categorical_to_interval(data=x_train)
                    pre_groups = copy.deepcopy(self.groups)
                # GridSearchCV parameters
                parameters_grid = {'max_leaf_nodes': [max_n_bins],
                                   'min_samples_leaf': [min_bin_size if min_bin_size > 1 else int(round(self.ds.samples[self.ds.train_name].shape[0] * min_bin_size, 0))],
                                   'max_depth': [max_depth]}
                # search for the best split
                # decision tree of split with pre-processed missings
                final_tree = self.fit_gridsearch(x_train, y_tran, parameters_grid, criterion, scoring)
                final_tree.fit(x_train[:, None], y_tran)
                self.groups = {}
                boundaries = sorted(self.get_tree_splits(final_tree))
                for i in range(len(boundaries) - 1):
                    self.groups[i] = [boundaries[i], boundaries[i + 1]]
                self.set_round_groups(with_acceptable_migration=True, verbose=verbose)
                if self.categorical_type:
                    self.categorical_recover(pre_groups)
            else:
                if min_bin_size > 1:
                    min_bin_size = min_bin_size / self.ds.samples[self.ds.train_name].shape[0]
                optb = optbinning.OptimalBinning(dtype='categorical' if self.categorical_type else 'numerical',
                                                 prebinning_method='cart', max_n_prebins=max_n_bins, min_prebin_size=min_bin_size,
                                                 monotonic_trend='auto_asc_desc' if monotonic else 'auto',
                                                 solver=solver, divergence=divergence)
                optb.fit(x_train, y_tran)
                if self.categorical_type:
                    self.groups = {g: sorted(list(val)) for g, val in enumerate(optb.splits)}
                else:
                    self.groups = {}
                    boundaries = [-np.inf] + list(optb.splits) + [np.inf]
                    for i in range(len(boundaries) - 1):
                        self.groups[i] = [boundaries[i], boundaries[i + 1]]
                    self.set_round_groups(with_acceptable_migration=True, verbose=verbose)

            if self.opposite_sign_to_others and not self.categorical_type:
                if df[df[self.feature] < 0].empty:
                    for group, vals in self.groups.items():
                        if isinstance(vals, list) and str(vals[0]) == '-inf':
                            self.groups[group][0] = 0
                            break
                if df[df[self.feature] >= 0].empty:
                    for group, vals in self.groups.items():
                        if isinstance(vals, list) and str(vals[1]) == 'inf':
                            self.groups[group][1] = 0
                            break
            if self.special_bins:
                for i, val in enumerate(self.special_bins.values()):
                    if not self.ds.samples[self.ds.train_name][self.ds.samples[self.ds.train_name][self.feature] == val].empty:
                        self.groups[-2 - i] = [val]

        if missings_to_process:
            self.groups[-1] = np.nan
        # WOE calculation
        self.groups = dict(sorted(self.groups.items()))
        self.set_groups(inplace=True)
        self.woe_fit(verbose=verbose, missings_to_process=missings_to_process)
        self.ds.samples[self.ds.train_name][add_suffix(self.feature)] = self.ds.samples[self.ds.train_name]['group'].map(self.woes)
        if to_history:
            self.to_history()
        if len(self.groups) != len(self.woes):
            if (-1 in self.woes and -1 not in self.groups) and pd.isnull(self.woes[-1]):
                del self.woes[-1]
            elif (-1 in self.groups and -1 not in self.woes) and pd.isnull(self.groups[-1]):
                del self.groups[-1]
            elif  verbose:
                print ('WARNING! Number of groups is not certain! We have', len(self.groups), 'groups and', len(self.woes), 'woes!')
        if verbose > 0:
            self.print_woe()

    def merge(self, groups_list, verbose=False):
        '''
        Merges two WOE intervals

        Parameters
        -----------
        groups_list: [group1, group2] - the groups to be merged
        '''

        # Checks for correctness of the groups to merge
        # existing groups
        for group in groups_list:
            if group not in self.groups:
                print(f'Group {group} is incorrect! Correct values are {list(self.groups)}')
                return None
        # only 2 groups per merge call
        if len(groups_list) != 2:
            print ('Please enter 2 groups')
            return None

        # only neighbouring groups for ordered features
        if not self.categorical_type:
            if -1 not in groups_list:
                if self.groups[groups_list[0]][0] != self.groups[groups_list[1]][1] and self.groups[groups_list[1]][0] != self.groups[groups_list[0]][1]:
                    print ('Please enter neighbouring groups.')
                    return None

        # merging groups in self.groups
        min_group = min(groups_list)
        max_group = max(groups_list)

        if not isinstance(self.groups[min_group], list):
            self.missing_group = max_group
            del self.groups[min_group]
        else:
            if self.categorical_type:
                self.groups[min_group] = self.groups[min_group] + self.groups[max_group]
            else:
                self.groups[min_group] = [min(self.groups[min_group][0], self.groups[max_group][0]), max(self.groups[min_group][1], self.groups[max_group][1])]

            if self.missing_group==max_group:
                self.missing_group=min_group
            elif self.missing_group>max_group:
                self.missing_group=self.missing_group-1

            shift_flag = False
            new_groups = copy.deepcopy(self.groups)
            for group in sorted([g for g in new_groups if not isinstance(g, str)]):
                if group == max_group:
                    shift_flag = True
                elif group > max_group:
                    shift_flag = True
                    self.groups[group - 1] = self.groups[group].copy()
            if shift_flag:
                del self.groups[max([g for g in self.groups if not isinstance(g, str)])]

        pre_woes_backup = list(self.ds.samples[self.ds.train_name][self.feature])
        self.fit(new_groups=False, verbose=verbose)
        gc.collect()
        self.ds.samples[self.ds.train_name][self.feature] = pre_woes_backup

    def to_history(self, dupl=True):
        '''
        Writes current state of self to history.
        '''
        current = {'groups':self.groups.copy(),
                   'woes': self.woes.copy(),
                   'missing_group': self.missing_group,
                   'others_woe':self.others_woe}
        if dupl or current not in self.history:
            self.history.append(current)

    def curr_iteration(self):
        try:
            curr_iteration = len(self.history) - 1 - self.history[::-1].index({'groups': self.groups,
                                                                               'woes': self.woes,
                                                                               'missing_group': self.missing_group,
                                                                               'others_woe': self.others_woe})
        except:
            self.to_history()
            curr_iteration = len(self.history) - 1
        return curr_iteration

    def insert_subgroup(self, group, new_group):
        '''
        Makes new group and calculates WOE for categorical features

        Parameters
        --------------
        new_group: a user-defined new group of values, consists of values from the group to split and the other values of the group will be separated.
        Only for categorical features. Example: group = [1, 2, 4, 6, 9], new_group = [1, 2, 9] => the two new groups will be [1, 2, 9], [4, 6].
        For the same result we could set new_group parameter = [4, 6]
        '''
        self.groups[max(self.groups) + 1] = new_group
        self.groups[group] = [i for i in self.groups[group] if i not in new_group]
        self.fit(new_groups=False)

    def insert_new_bound(self, group, add_bound):
        '''
        Inserts new bound and calculates WOE for interval features

        Parameters
        ------------
        group: the group to insert the new bound into
        add_bound: the new bound to insert
        '''
        new_group_num = group + 1
        tmp_groups = copy.deepcopy(self.groups)
        tmp_woes = copy.deepcopy(self.woes)
        for g in sorted([g for g in self.groups if not isinstance(g, str)]):
            if g > new_group_num:
                tmp_groups[g] = self.groups[g - 1].copy()
                tmp_woes[g] = self.woes[g - 1]
        tmp_groups[max(self.groups) + 1] = self.groups[max(self.groups)].copy()
        tmp_woes[max(self.woes) + 1] = self.woes[max(self.woes)]
        if self.missing_group > group:
            self.missing_group = self.missing_group+1
        self.groups = tmp_groups
        self.woes = tmp_woes
        self.groups[new_group_num] = [add_bound, self.groups[group][1]]
        self.groups[group][1] = add_bound
        self.fit(new_groups=False)

    def split(self, group=None, to_add=None, min_bin_size=0.05, criterion='entropy', scoring='neg_log_loss'):
        '''
        Splits a WOE interval into two.

        Parameters
        -----------
        group: a group to split, integer
        to_add: in case of interval - a user-defined bound for the split (the intermediate bound of the interval), only for ordered features; in case of categorical -  a user-defined new group of values, consists of values from the group to split and the other values of the group will be separated. Only for categorical features. Example: group = [1, 2, 4, 6, 9], new_group = [1, 2, 9] => the two new groups will be [1, 2, 9], [4, 6]. For the same result we could set new_group parameter = [4, 6]
        '''
        if group is None:
            if self.categorical_type:
                for g in self.groups:
                    if to_add[0] in self.groups[g]:
                        group = g
                        break
            else:
                for g, v in self.groups.items():
                    if to_add >= v[0] and to_add < v[1]:
                        group = g
                        break

        if group is None or group == -1:
            print('Invalid group!')
            return None

        if isinstance(to_add, int) or isinstance(to_add, float):
            if self.categorical_type:
                print ('The feature is categorical so to_add should be a list of values for the new group. Good luck!')
                return None
            elif not (to_add >= self.groups[group][0] and to_add < self.groups[group][1]):
                print('New bound is out-of-range for the specified group. Bye-bye.')
                return None
            else:
                self.insert_new_bound(group, to_add)
        elif isinstance(to_add, list) or isinstance(to_add, np.ndarray):
            if not self.categorical_type:
                print ('The feature is not categorical so to_add must be a float. Good luck!')
                return None
            else:
                for n in to_add:
                    if n not in self.groups[group]:
                        print('Invalid new_group!')
                        return None
                if self.groups[group] == to_add:
                    print ('Error: new_group contains all the values of group', group)
                    return None
                else:
                    self.insert_subgroup(group, to_add)
                    gc.collect()

        # if no pre-defined bounds or groups
        else:
            print ('Splitting started! Feature', self.feature, 'group:', group)
            df = self.ds.samples[self.ds.train_name]
            if self.categorical_type:
                df = df[df[self.feature].isin(self.groups[group])].copy()
            else:
                df = df[(df[self.feature] >= self.groups[group][0]) & (df[self.feature] < self.groups[group][1])].copy()

            if df[self.ds.target].nunique() > 1:
                parameters_grid = {'min_samples_leaf': [min_bin_size if min_bin_size > 1 else int(round(self.ds.samples[self.ds.train_name].shape[0] * min_bin_size, 0))],
                                   'max_depth' : [1]}

                tmp_categorical = self.categorical_type
                self.categorical_type = ''
                # optimal split
                try:
                    final_tree = self.fit_gridsearch(df[self.feature], df[self.ds.target], parameters_grid, criterion, scoring)
                    final_tree.fit(df[self.feature][:, None], df[self.ds.target])
                except Exception:
                    print ('Fitting with cross-validation failed! Possible cause: too few representatives of one of the target classes.')
                    print ('Try setting the bound yourself')
                    self.categorical_type = tmp_categorical
                    return None

                tree_splits = self.get_tree_splits(final_tree)
                self.categorical_type = tmp_categorical
                tree_splits = [x for x in tree_splits if x not in [-np.inf, np.inf]]
                if len(tree_splits) == 0:
                    print ('No good binning found. Try setting the bound yourself')
                    return None
                else:
                    add_bound = sorted(tree_splits)[0]
                    print ('Additional bound ', add_bound)
                    #adding the new bound (woe for categorical) to groups
                    if self.categorical_type:
                        # find group by woe bound...
                        # since in self.ds.samples[self.ds.train_name][feature] we have woes calculated for each categorical value...
                        new_group = list(self.ds.samples[self.ds.train_name][(self.ds.samples[self.ds.train_name][self.feature] < add_bound) & (self.ds.samples[self.ds.train_name][self.feature].isin(self.groups[group]))][self.feature].drop_duplicates())
                        print ('new_group:', new_group)
                        self.insert_subgroup(group, new_group)
                    else:
                        self.insert_new_bound(group, add_bound)
                    gc.collect()
            else:
                print(f'All observations in the specified group have the same target value {df[self.ds.target].unique()[0]}')

    def merge_by_woe(self, woe_threshold=0.05, verbose=False):
        '''
        Merges all groups, close by WOE (for interval features only neighboring groups and missing group are checked)

        Parameters
        -----------
        woe_threshold: if woe difference between groups (neighboring groups for interval) is less then this threshold, then they are to be merged
        '''
        if verbose > 0:
            print('\n----------------- Merging by WOE -----------------')

        for _ in range(len(self.groups.copy())):
            groups_stat = self.calc_groups_stat()
            groups_stat = groups_stat[groups_stat.index >= 0].reset_index()
            if self.categorical_type:
                groups_stat = groups_stat.sort_values(by=['woe'])
            groups_stat['woe_diff'] = (groups_stat['woe'] - groups_stat['woe'].shift(-1)).abs()
            groups_stat['merge_to'] = groups_stat['group'].shift(-1)
            groups_stat = groups_stat[(groups_stat['woe_diff'] < woe_threshold) & (groups_stat['woe_diff'] == groups_stat['woe_diff'].min())]

            if not groups_stat.empty:
                merge_groups = list(groups_stat[['group', 'merge_to']].astype('int').values[0])
                if verbose > 0:
                    print(f'\nMerging of two groups close by WOE: {merge_groups}')
                self.merge(merge_groups, verbose=verbose)
            else:
                if verbose > 0:
                    print('No groups with close WOE.')
                    print('...Done')
                break

    def merge_by_size(self, target_threshold=5, size_threshold=100, verbose=False):
        '''
        Merges small groups (by target or size) to the closest by WoE (for interval features only neighboring groups and missing group are checked)

        Parameters
        -----------
        target_threshold: min number of targets for group to not be considered small
        size_threshold: min number of observations for group to not be considered small
        '''
        if verbose > 0:
            print('\n---------------- Merging by size -----------------')

        for _ in range(len(self.groups.copy())):
            groups_stat = self.calc_groups_stat()
            groups_stat = groups_stat[groups_stat.index >= 0].reset_index()
            if self.categorical_type:
                groups_stat = groups_stat.sort_values(by=['woe'])
            if target_threshold < 1:
                groups_stat['n1'] = groups_stat['n1'] / groups_stat['n']
            groups_stat['small'] = (groups_stat['n1'] < target_threshold) | (groups_stat['n'] < size_threshold)
            groups_stat['woe_diff1'] = np.where(groups_stat['small'], (groups_stat['woe'] - groups_stat['woe'].shift(1)).abs(), np.nan)
            groups_stat['woe_diff2'] = np.where(groups_stat['small'], (groups_stat['woe'] - groups_stat['woe'].shift(-1)).abs(), np.nan)
            groups_stat['woe_diff'] = groups_stat[['woe_diff1', 'woe_diff2']].min(axis=1)
            groups_stat['merge_to'] = np.where(groups_stat['woe_diff'] != groups_stat['woe_diff'].min(), np.nan,
                                               np.where(groups_stat['woe_diff'] == groups_stat['woe_diff1'], groups_stat['group'].shift(1),
                                                        np.where(groups_stat['woe_diff'] == groups_stat['woe_diff2'], groups_stat['group'].shift(-1), np.nan)))
            groups_stat = groups_stat.dropna(subset=['merge_to'])

            if not groups_stat.empty:
                merge_groups = list(groups_stat[['group', 'merge_to']].astype('int').values[0])
                if verbose > 0:
                    print(f'\nMerging small group {merge_groups[0]} (by target or size) to the closest by WoE group {merge_groups[1]}')
                self.merge(merge_groups, verbose=verbose)
            else:
                if verbose > 0:
                    print('No small groups.')
                    print('...Done')
                break

    def transform(self):
        '''
        Transforms a Data object according to WOE parameters fitted. Can be used only after .fit().

        Parameters
        ------------
        data: Data object to transform

        Returns
        ----------
        transformed Data object
        '''
        if self.ds.samples is not None:
            f_WOE = add_suffix(self.feature)
            for name, sample in self.ds.samples.items():
                self.ds.samples[name][f_WOE] = self.set_avg_woes(data=(self.ds.samples[name][self.feature]))
            if self.ds.bootstrap_base is not None:
                self.ds.bootstrap_base[f_WOE] = self.set_avg_woes(data=(self.ds.bootstrap_base[self.feature]))
            self.ds.features = [f_WOE]
        return self.ds

    def export_scorecard(self, iteration=None, full=True):
        '''
        Transforms self.groups, self.categorical and self.missing_group to dataframe.

        Returns
        ----------
        dataframe with binning information
        '''
        # searching for WOE for each interval of values
        if self.ds is not None:
            scorecard = self.calc_groups_stat().reset_index()
        else:
            scorecard = pd.DataFrame.from_dict(self.woes, orient='index').reset_index().set_axis(['group', 'woe'], axis=1)
            scorecard['n'] = np.nan
            scorecard['n0'] = np.nan
            scorecard['n1'] = np.nan
        scorecard['values'] = scorecard['group'].map(self.groups)
        scorecard = scorecard[scorecard['group'] != 404]
        scorecard = pd.concat([scorecard, pd.DataFrame({'group': 'others', 'values': 'all others', 'woe': self.others_woe}, index=[0])])
        scorecard['categorical_type'] = self.categorical_type
        scorecard['missing'] = np.where(scorecard['group'] == self.missing_group, 1, 0)
        scorecard['feature'] = self.feature
        scorecard = scorecard[['feature', 'categorical_type', 'group', 'values', 'woe', 'missing', 'n', 'n0', 'n1']]
        scorecard['target_rate'] = scorecard['n1'] / scorecard['n']
        scorecard['sample_part'] = scorecard['n'] / scorecard['n'].sum()
        scorecard['n0_part'] = scorecard['n0'] / scorecard['n0'].sum()
        scorecard['n1_part'] = scorecard['n1'] / scorecard['n1'].sum()
        scorecard['iteration'] = iteration if iteration is not None else len(self.history)
        scorecard[['target_rate', 'sample_part', 'n0_part', 'n1_part']] = scorecard[['target_rate', 'sample_part', 'n0_part', 'n1_part']].round(self.round_woe)
        for i, group in enumerate(self.special_bins.keys()):
            scorecard.loc[scorecard['group'] == -2 - i, 'group'] = group
        if not full:
            scorecard.drop(['n', 'n0', 'n1', 'n0_part', 'n1_part', 'iteration'], axis=1, inplace=True)
        return scorecard

    def rollback(self, iteration=None):
        '''
        Rolls back the last operation.

        Parameters
        -----------
        iteration: number of groups iteration to return to (if None, then rollback to the previous iteration)
        '''
        if iteration is None:
            iteration = -1
        if self.history and iteration < len(self.history):
            self.groups = self.history[iteration]['groups']
            self.woes = self.history[iteration]['woes']
            self.missing_group = self.history[iteration]['missing_group']
            self.others_woe = self.history[iteration]['others_woe']
            self.set_groups(inplace=True)
            self.ds.samples[self.ds.train_name][add_suffix(self.feature)] = self.ds.samples[self.ds.train_name]['group'].map(self.woes)
            self.is_active = True
            gc.collect()
        else:
            print ('Sorry, no changes detected or iteration found. Nothing to rollback.')
            return None

    def check_values(self, s):
        '''
        Checks if any string element of the list contains comma. This method is used in parsing imported dataframes with groups, borders and woes.

        Returns
        --------
        False if there is a comma
        '''
        quotes = s.count("'")
        if quotes > 0:
            commas = s.count(',')
            if commas != (quotes/2)-1:
                return False
        return True

    def str_to_list(self, s):
        '''
        Parses ['values'] from a dataframe constructed by self.groups_to_dataframe().
        '''
        s = str(s)

        if pd.isnull(s) or s == '[nan]':
            return np.nan
        if self.check_values(s):
            v = (re.split('[\'|"]? *, *[\'|"]?', s[1:-1]))
            if v[0][0] in ("'", '"'):
                v[0]=v[0][1:]
            if v[-1][-1] in ("'", '"'):
                v[-1]=v[-1][:-1]
            return [float(x) if (x[-3:] == 'inf' or (min([y.isdigit() for y in x.split('.')]) and x.count('.') < 2)) else (x if x!='' else np.nan) for x in v]
        else:
            print ('Error in string', s, '! Delete commas from feature values!')
            return None

    def import_scorecard(self, scorecard, verbose=True, fit_flag=False):
        '''
        Sets self.groups, self.categorical_type and self.missing_group values from dataframe and calculates woe (by fit).

        Parameters
        ----------
        scorecard: a DataFrame with 'categorical_type', 'group', 'values' and 'missing' fields
        verbose: should bins and woes be printed or not
        fit_flag: should woes be calculated or taken from input dataframe
        '''
        if 'woe' not in scorecard:
            fit_flag = True
        if 'iteration' not in scorecard:
            scorecard['iteration'] = 0
        self.is_active = True
        for iter, woe_df in scorecard.groupby('iteration'):
            try:
                self.others_woe = woe_df[woe_df['group'] == 'others']['woe'].iloc[0]
            except:
                self.others_woe = np.nan
            woe_df = woe_df[~woe_df['group'].isin(['others'])]
            if 'categorical_type' in woe_df:
                self.categorical_type = list(woe_df['categorical_type'].fillna(''))[0]
            elif 'categorical' in woe_df:
                self.categorical_type = 'object' if list(woe_df['categorical'])[0] else ''
            values = list(woe_df['values'])
            to_convert = False
            for v in values:
                if isinstance(v, str):
                    to_convert = True
            if self.categorical_type:
                if to_convert:
                    values_corrected = []
                    for v in values:
                        if isinstance(v, str):
                            if self.categorical_type == 'object':
                                v_ = re.split('[\'|"] *, *[\'|"]', v[1:-1])
                            else:
                                v_ = re.split(' *, *', v[1:-1])
                            if v_[0][0] in ("'", '"'):
                                v_[0] = v_[0][1:]
                            if v_[-1][-1] in ("'", '"'):
                                v_[-1] = v_[-1][:-1]
                        else:
                            v_ = v
                        values_corrected.append(np.array(v_).astype(self.categorical_type).tolist())
                    values = values_corrected.copy()
            else:
                if to_convert:
                    values_corrected = []
                    for v in values:
                        if str(v)[0] == '[':
                            values_corrected.append(
                                [float(x) if x != '' else np.nan for x in str(v)[1:-1].replace(" ", "").split(',')])
                        else:
                            values_corrected.append(float(v))
                    values = values_corrected.copy()
            woe_df['values'] = values
            if not self.special_bins:
                self.special_bins = {g: woe_df[woe_df['group'] == g]['values'].values[0] for g in woe_df['group'].values[::-1]
                                     if isinstance(g, str) and isinstance(woe_df[woe_df['group'] == g]['values'].values[0], list)}
            for i, g in enumerate(self.special_bins):
                woe_df.loc[woe_df['group'] == g, 'group'] = -2 - i
            self.groups = woe_df.set_index('group')['values'].to_dict()
            try:
                self.missing_group = woe_df[woe_df['missing'] == 1]['group'].values[0]
                if self.missing_group == -1:
                    self.groups[-1] = np.nan
            except:
                self.missing_group = -1
            self.groups = dict(sorted(self.groups.items()))
            if fit_flag and iter == scorecard['iteration'].max():
                self.fit(new_groups=False)
            else:
                self.woes = woe_df.set_index('group')['woe'].to_dict()
                if self.ds.samples is not None and iter == scorecard['iteration'].max():
                    self.set_groups(inplace=True)
                    self.ds.samples[self.ds.train_name][add_suffix(self.feature)] = self.ds.samples[self.ds.train_name]['group'].map(self.woes)
                    if verbose > 0:
                        self.print_woe()
                self.to_history(dupl=False)


class FeatureCross:
    '''
    ВОЕ-трансформация для кросс-переменных. Для каждой переменной первого уровня должэен быть создан свой эксземпляр
    '''

    def __init__(self, feature_woe):
        """
        :param feature_woe: объект FeatureWOE переменной первого уровня
        """
        self.feature_woe = copy.deepcopy(feature_woe)
        self.cross = {}

    def add_f2_features(self, feature_woes):
        group_idx = {}
        if self.feature_woe.ds.samples is not None:
            for sample in self.feature_woe.ds.samples:
                if sample != self.feature_woe.ds.train_name:
                    df_s = self.feature_woe.ds.samples[sample].copy()
                    df_s['group'] = self.feature_woe.set_groups(data=df_s[self.feature_woe.feature])
                else:
                    df_s = self.feature_woe.ds.samples[sample]
                group_idx[sample] = {group: [df_s.index.get_loc(idx) for idx in df_sg.index] for group, df_sg in
                                     df_s.groupby('group')}
        if self.feature_woe.ds.bootstrap_base is not None:
            bs = self.feature_woe.ds.bootstrap_base
            bs['group'] = self.feature_woe.set_groups(data=bs[self.feature_woe.feature])
            bs_group_idx = {group: {bs.index.get_loc(idx) for idx in tmp.index} for group, tmp in bs.groupby('group')}
            bootstrap = {
                group: [list(set(idx) & bs_group_idx[group]) for idx in self.feature_woe.ds.bootstrap] for
                group in bs_group_idx}
        for fw in feature_woes:
            self.cross[fw.feature] = {}
            for group in self.feature_woe.groups:
                if fw.ds.samples is not None:
                    samples = {name: sample.iloc[group_idx[name][group]] for name, sample in fw.ds.samples.items() if group in group_idx[name]}
                else:
                    samples = None
                fw2 = FeatureWOE(DataSamples(samples=samples, target=fw.ds.target, features=fw.ds.features, cat_columns=fw.ds.cat_columns, n_jobs=1, random_state=fw.ds.random_state),
                                 fw.feature, round_digits=fw.round_digits, round_woe=fw.round_woe,
                          rounding_migration_coef=fw.rounding_migration_coef, simple=fw.simple, n_folds=fw.n_folds,
                          woe_adjust=fw.woe_adjust, alpha=fw.alpha, alpha_range=fw.alpha_range,
                          alpha_scoring=fw.alpha_scoring, alpha_best_criterion=fw.alpha_best_criterion,
                          missing_process=fw.missing_process, missing_min_part=fw.missing_min_part,
                          others_process=fw.others_process, opposite_sign_to_others=fw.opposite_sign_to_others,
                          special_bins=fw.special_bins)
                if fw.ds.bootstrap_base is not None and self.feature_woe.ds.bootstrap_base is not None:
                    fw2.ds.bootstrap_base = fw.ds.bootstrap_base
                    fw2.ds.bootstrap = bootstrap[group]
                self.cross[fw.feature][group] = fw2

    def fit(self, cross_features=None, new_groups=True, verbose=False, method='tree', max_n_bins=10, min_bin_size=0.05,
            criterion='entropy', scoring='neg_log_loss', max_depth=None, monotonic=False, solver='cp', divergence='iv'):
        if cross_features is None:
            cross_features = list(self.cross)
        for f2 in cross_features:
            for group in self.cross[f2]:
                self.cross[f2][group].fit(new_groups=new_groups, to_history=True, verbose=verbose, method=method, max_n_bins=max_n_bins,
                     min_bin_size=min_bin_size, monotonic=monotonic, criterion=criterion, scoring=scoring, max_depth=max_depth, solver=solver, divergence=divergence)
            self.calc_simple_woes(f2)
            if verbose:
                self.print_woe(f2)

    def auto_fit(self, cross_features=None, verbose=True, method='tree', max_n_bins=10, min_bin_size=0.05,
                 criterion='entropy', scoring='neg_log_loss', max_depth=None, solver='cp', divergence='iv',
                 WOEM_on=True, WOEM_woe_threshold=0.05, WOEM_with_missing=False,
                 SM_on=True, SM_target_threshold=5, SM_size_threshold=100,
                 BL_on=True, BL_allow_Vlogic_to_increase_gini=100,
                 G_on=True, G_gini_threshold=5, G_with_test=False, G_gini_decrease_threshold=0.2,
                 G_gini_increase_restrict=True,
                 WOEO_on=False, WOEO_dr_threshold=0.01, WOEO_correct_threshold=0.85, WOEO_miss_is_incorrect=True,
                 WOEO_with_test=False
                 ):
        auto_fit_parameters = {}
        for f in ['method', 'max_n_bins', 'min_bin_size', 'criterion', 'scoring', 'max_depth', 'solver', 'divergence',
                  'WOEM_on', 'WOEM_woe_threshold', 'WOEM_with_missing', 'SM_on', 'SM_target_threshold',
                  'SM_size_threshold',
                  'G_on', 'G_gini_threshold', 'G_gini_decrease_threshold', 'G_gini_increase_restrict', 'G_with_test',
                  'WOEO_on', 'WOEO_dr_threshold', 'WOEO_correct_threshold', 'WOEO_miss_is_incorrect', 'WOEO_with_test',
                  'BL_on', 'BL_allow_Vlogic_to_increase_gini']:
            auto_fit_parameters[f] = eval(f)
        if cross_features is None:
            cross_features = list(self.cross)
        if not cross_features:
            return {}, {}
        check_dfs = {f2: {} for f2 in cross_features}
        result = {f2: False for f2 in cross_features}
        for f2 in cross_features:
            res_group = {}
            for group in self.cross[f2]:
                res_group[group], check_dfs[f2][group] = self.cross[f2][group].auto_fit(verbose=-1, **auto_fit_parameters)
                for i in range(5):
                    if not check_dfs[f2][group][i].empty:
                        check_dfs[f2][group][i] = check_dfs[f2][group][i][check_dfs[f2][group][i]['iteration'] == check_dfs[f2][group][i]['iteration'].max()].drop(['iteration'], axis=1)
                        check_dfs[f2][group][i].insert(loc=0, column='bin1', value=str(self.feature_woe.groups[group]))
                if not res_group[group]:
                    if self.cross[f2][group].categorical_type:
                        values = self.cross[f2][group].ds.samples[self.cross[f2][group].ds.train_name][self.cross[f2][group].feature].dropna().unique().tolist()
                        try:
                            self.cross[f2][group].groups = {0: sorted(values)}
                        except:
                            self.cross[f2][group].groups = {0: values}
                    else:
                        self.cross[f2][group].groups = {0: [-np.inf, np.inf]}
                    self.cross[f2][group].fit(new_groups=False)
            result[f2] = any(list(res_group.values()))
            self.calc_simple_woes(f2)
            if verbose:
                self.print_woe(f2)
        check_dfs = [pd.concat([check_dfs[f2][group][i] for f2 in cross_features for group in check_dfs[f2]]) for i in range(5)] + [self.export_scorecard()]
        return result, check_dfs

    def calc_simple_woes(self, f2):
        to_calc = self.feature_woe.ds.samples[self.feature_woe.ds.train_name].merge(pd.concat(
            [fw.ds.samples[fw.ds.train_name]['group'] for g, fw in
             self.cross[f2].items()]), left_index=True, right_index=True)
        to_calc['group'] = '[' + to_calc['group_x'].astype('str') + ', ' + to_calc['group_y'].astype('str') + ']'
        for g, woe in self.feature_woe.calc_simple_woes(to_calc).items():
            g_l = json.loads(g)
            self.cross[f2][g_l[0]].woes[g_l[1]] = woe

    def calc_groups_stat(self, f2):
        to_calc = self.feature_woe.ds.samples[self.feature_woe.ds.train_name].merge(pd.concat(
            [fw.ds.samples[fw.ds.train_name]['group'] for g, fw in
             self.cross[f2].items()]), left_index=True, right_index=True)
        to_calc['group'] = '[' + to_calc['group_x'].astype('str') + ', ' + to_calc['group_y'].astype('str') + ']'
        groups_stat = self.feature_woe.calc_groups_stat(to_calc)
        groups_stat['woe'] = groups_stat.index.map(self.get_woes(f2))
        return groups_stat

    def print_woe(self, f2):
        '''
        Prints WOE parameters in a standard and convenient way
        '''
        groups_stat = self.calc_groups_stat(f2).reset_index()
        groups_stat[['n', 'n1']] = groups_stat[['n', 'n1']].round(0).astype('int')
        groups_stat['values'] = groups_stat['group'].map(self.get_2values(f2)).astype('str')
        '''
        if self.missing_group != -1 and not self.categorical_type:
            groups_stat.loc[groups_stat['group'] == self.missing_group, 'values'] += ' + missings'
        for i, group in enumerate(self.special_bins.keys()):
            groups_stat.loc[groups_stat['group'] == -2 - i, 'group'] = group
        '''
        print(f'\nCurrent binning for {cross_name(self.feature_woe.feature, f2)}:')
        print(groups_stat[['group', 'values', 'woe', 'n', 'n1']].to_string(index=False))
        print()

    def get_transform_func(self, f2, start=''):
        groups1 = self.feature_woe.sorted_groups()
        s = start
        i = 0
        for g1 in groups1:
            fw = self.cross[f2][g1]
            groups2_list = [g for g in fw.sorted_groups() if g in fw.woes]
            g1_condition_code = self.feature_woe.get_group_condition_code(g1)
            for g2 in groups2_list:
                g2_condition_code = fw.get_group_condition_code(g2)
                s += f"np.where(({g1_condition_code}) & ({g2_condition_code}), {fw.woes[g2]}, \n" + " " * (len(start) + 9 * (i + 1))
                i += 1
        s += f'{self.feature_woe.others_woe}{")" * i}'.replace('nan', 'np.nan')
        return s

    def set_avg_woes(self, data, f2):
        '''
        Replaces all values of a feature to related WOE

        Parameters
        -----------
        data: DataFrame, containing initial values of features

        Returns
        -----------
        a Series of WOE-transformed feature values
        '''
        condlist = []
        cond_woes = []
        condlist_1, groups_1 = self.feature_woe.get_condlist(data[self.feature_woe.feature], self.feature_woe.groups)
        for cond, g in zip(condlist_1, groups_1):
            condlist_2, groups_2 = self.cross[f2][g].get_condlist(data[f2], self.cross[f2][g].groups)
            condlist += [c & cond for c in condlist_2]
            cond_woes += [self.cross[f2][g].woes[g2] for g2 in groups_2]
        return np.select(condlist, cond_woes, self.feature_woe.others_woe)

    def set_groups(self, data, f2):
        '''
        Replaces all values of a feature to related WOE

        Parameters
        -----------
        data: DataFrame, containing initial values of features

        Returns
        -----------
        a Series of WOE-transformed feature values
        '''
        condlist = []
        cond_groups = []
        condlist_1, groups_1 = self.feature_woe.get_condlist(data[self.feature_woe.feature], self.feature_woe.groups)
        for cond, g in zip(condlist_1, groups_1):
            condlist_2, groups_2 = self.cross[f2][g].get_condlist(data[f2], self.cross[f2][g].groups)
            condlist += [c & cond for c in condlist_2]
            cond_groups += [str([g, g2]) for g2 in groups_2]
        return np.select(condlist, cond_groups, 404)

    def transform(self, features=None):
        '''
        Transforms a Data object according to WOE parameters fitted. Can be used only after .fit().

        Parameters
        ------------
        data: Data object to transform

        Returns
        ----------
        transformed Data object
        '''
        if features is None:
            features = list(self.cross)
        if self.feature_woe.ds.samples is not None:
            new_features = []
            for f2 in features:
                f_WOE = add_suffix(cross_name(self.feature_woe.feature, f2))
                for name, sample in self.feature_woe.ds.samples.items():
                    data = self.feature_woe.ds.samples[name].merge(pd.concat([fw.ds.samples[name][f2] for g, fw in self.cross[f2].items() if name in fw.ds.samples]), how='left', left_index=True, right_index=True)
                    self.feature_woe.ds.samples[name][f_WOE] = self.set_avg_woes(data, f2)
                if self.feature_woe.ds.bootstrap_base is not None:
                    data = self.feature_woe.ds.bootstrap_base.merge(list(self.cross[f2].values())[0].ds.bootstrap_base[f2], how='left', left_index=True, right_index=True)
                    self.feature_woe.ds.bootstrap_base[f_WOE] = self.set_avg_woes(data, f2)
                new_features.append(f_WOE)
            self.feature_woe.ds.features = new_features
        return self.feature_woe.ds

    def export_scorecard(self, features=None, full=True):
        '''
        Transforms self.groups, self.categorical and self.missing_group to dataframe.

        Returns
        ----------
        dataframe with binning information
        '''
        if features is None:
            features = self.cross
        features = [f for f in features if f in self.cross]
        dfs = []
        for f2 in features:
            feature = f'cross_{self.feature_woe.feature}&{f2}'
            dfs_group = []
            for group in self.cross[f2]:
                df = self.cross[f2][group].export_scorecard(full=full)
                df = df[df['group'] != 'others']
                df['group'] = '[' + str(group) + ', ' + df['group'].astype('str') + ']'
                df['values'] = str(self.feature_woe.groups[group]) + ' & ' + df['values'].astype('str')
                df['missing'] = '[' + str(int(self.feature_woe.missing_group == group)) + ', ' + df['missing'].astype('str') + ']'
                df['missing'].loc[df['missing'] == '[0, 0]'] = 0
                dfs_group.append(df)
            df = pd.concat(dfs_group).reset_index(drop=True)
            df['feature'] = feature
            categorical_type = str([self.feature_woe.categorical_type, df['categorical_type'].values[0]]).replace("'","").replace("[, ]","")
            df['categorical_type'] = categorical_type
            if full:
                df['sample_part'] = df['n'] / df['n'].sum()
                df['n0_part'] = df['n0'] / df['n0'].sum()
                df['n1_part'] = df['n1'] / df['n1'].sum()
                df[['sample_part', 'n0_part', 'n1_part']] = df[['sample_part', 'n0_part', 'n1_part']].round(self.feature_woe.round_woe)
            dfs.append(df)
            dfs.append(pd.DataFrame({'feature': feature, 'categorical_type': categorical_type, 'group': 'others', 'values': 'all others', 'woe': self.feature_woe.others_woe, 'missing': 0}, index=[0]))
        if dfs:
            scorecard = pd.concat(dfs).reset_index(drop=True)
        else:
            scorecard = pd.DataFrame()
        return scorecard

    def import_scorecard(self, scorecard, verbose=True, fit_flag=False):
        '''
        Sets self.groups, self.categorical_type and self.missing_group values from dataframe and calculates woe (by fit).

        Parameters
        ----------
        scorecard: a DataFrame with 'categorical_type', 'group', 'values' and 'missing' fields
        verbose: should bins and woes be printed or not
        fit_flag: should woes be calculated or taken from input dataframe
        '''
        for f2, scorecard_f2 in scorecard.groupby('f2'):
            if scorecard_f2[scorecard_f2['group1'] != 'others'].empty:
                continue
            for group, scorecard_f2_group in scorecard_f2.groupby('group1'):
                if group != 'others':
                    self.cross[f2][group].import_scorecard(scorecard=scorecard_f2_group[['f2', 'categorical_type2', 'group2', 'values2', 'woe', 'missing2']]\
                                            .set_axis(['feature', 'categorical_type', 'group', 'values', 'woe', 'missing'], axis=1),
                                            verbose=False, fit_flag=False)

                else:
                    self.feature_woe.others_woe = scorecard_f2_group['woe'].values[0]
            if fit_flag:
                self.calc_simple_woes(f2)
            if verbose:
                self.print_woe(f2)

    def get_woes(self, f2):
        return {str([g, g2]): woe for g, fw in self.cross[f2].items() for g2, woe in fw.woes.items()}

    def get_values(self, f2):
        return {str([g, g2]): value for g, fw in self.cross[f2].items() for g2, value in fw.groups.items()}

    def get_2values(self, f2):
        return {str([g, g2]): str(self.feature_woe.groups[g]) + ' & ' + str(value) for g, fw in self.cross[f2].items() for g2, value in fw.groups.items()}